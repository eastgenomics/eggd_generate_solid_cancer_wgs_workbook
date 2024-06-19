import argparse
import os
import re
import subprocess
import urllib.request
from openpyxl import drawing
from openpyxl.styles import Alignment, Border, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import pandas as pd
from bs4 import BeautifulSoup
from PIL import Image


# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)
DEFAULT_FONT.name = "Calibri"


class excel:
    """
    Functions for wrangling input csv files, ref files and html files and
    writing output xlsx file

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        writer object for writing Excel data to file
    workbook : openpyxl.workbook.workbook.Workbook
        openpyxl workbook object for interacting with per-sheet writing and
        formatting of output Excel file

    Outputs
    -------
    {args.output}.xlsx : file
        Excel file with variants and structural variants
    """

    def __init__(self) -> None:
        self.args = self.parse_args()
        basename = os.path.basename(self.args.variant)
        self.output_file = basename.rsplit("-", 1)[0] + ".xlsx"
        self.writer = pd.ExcelWriter(self.output_file, engine="openpyxl")
        self.workbook = self.writer.book

    def parse_args(self) -> argparse.Namespace:
        """
        Parse command line arguments

        Returns
        -------
        args : Namespace
            Namespace of passed command line argument inputs
        """
        parser = argparse.ArgumentParser()
        parser.add_argument("-html", required=True, help="html input")
        parser.add_argument(
            "--variant", "-v", required=True, help="variant csv file"
        )
        parser.add_argument(
            "--SV", "-sv", required=True, help="structural variant csv file"
        )
        parser.add_argument(
            "--hotspots", "-hs", required=True, help="hotspots ref file"
        )
        parser.add_argument(
            "--refgene_group",
            "-rgg",
            required=True,
            help="refgene group ref file",
        )
        parser.add_argument(
            "--clinvar", "-c", required=True, help="clinvar ref file"
        )
        return parser.parse_args()

    def generate(self) -> None:
        """
        Calls all methods in excel() to generate output xlsx
        """
        self.download_html_img()
        self.write_sheets()
        self.workbook.save(self.output_file)
        print("Done!")

    def download_html_img(self) -> None:
        """
        get the image links from html file
        """
        soup = self.get_soup()
        n = 1
        for link in soup.findAll("img"):
            img_link = link.get("src")
            self.download_image(img_link, "./", f"figure_{n}")
            n = n + 1
        self.crop_img("figure_2.jpg", 600, 600, 2400, 2400)

    def download_image(self, url, file_path, file_name) -> None:
        """
        Download the img from html links
        Parameters
        ----------
        url
        str for file path
        str for file name
        """
        full_path = file_path + file_name + ".jpg"
        urllib.request.urlretrieve(url, full_path)

    def crop_img(self, img_to_crop, left, top, right, bottom) -> None:
        """
        crop the image
        Parameters
        ---------
        str - img file name
        int - left margin to crop
        int - top margin to crop
        int - right margin to crop
        int - bottom margin to crop
        """
        im = Image.open(img_to_crop)
        im1 = im.crop((left, top, right, bottom))
        im1.save("cropped_" + img_to_crop)

    def read_html_tables(self, table_num) -> list:
        """
        get the tables from html file

        Parameters
        ----------
        int - table number

        Returns
        -------
        list of html table
        """
        soup = self.get_soup()
        tables = soup.findAll("table")
        info = tables[table_num]
        datasets = self.extract_data_from_html_table(info)

        return datasets

    def extract_data_from_html_table(self, table_info) -> list:
        """
        strip html table and return as list
        Parameters
        ----------
        bs4.element.Tag for html table table

        Returns
        -------
        list of html table
        """
        headings = [
            th.get_text() for th in table_info.find("tr").find_all("th")
        ]
        datasets = []
        for row in table_info.find_all("tr")[1:]:
            dataset = dict(
                zip(headings, (td.get_text() for td in row.find_all("td")))
            )
            datasets.append(dataset)
        return datasets

    def get_soup(self) -> BeautifulSoup:
        """
        get Beautiful soup obj from html

        Returns
        -------
        Beautiful soup object
        """
        url = self.args.html
        page = open(url)
        soup = BeautifulSoup(page.read())
        return soup

    def get_tmb(self) -> BeautifulSoup:
        """
        get tumor mutation burden (tmb) from html

        Returns
        -------
        bs4.element.NavigableString for tmb
        """
        soup = self.get_soup()
        pattern = re.compile(
            (
                "Total number of somatic non-synonymous"
                " small variants per megabase"
            )
        )
        tmb = soup.find("b", text=pattern).next_sibling
        return tmb

    def write_sheets(self) -> None:
        """
        Write sheets to xlsx file
        """
        print("Writing sheets")
        self.soc = self.workbook.create_sheet("SOC")
        self.write_soc()
        self.QC = self.workbook.create_sheet("QC")
        self.write_QC()
        self.plot = self.workbook.create_sheet("Plot")
        self.write_plot()
        self.signatures = self.workbook.create_sheet("Signatures")
        self.write_signatures()
        self.germline = self.workbook.create_sheet("Germline")
        self.write_germline()
        self.write_SNV()
        self.write_gain_loss()
        self.write_SV()
        self.summary = self.workbook.create_sheet("Summary")
        self.write_summary()
        self.write_refgene()

    def set_col_width(self, cell_width, sheet) -> None:
        """
        set the column width for given col in given sheet
        Parameters
        ----------
        tuple for cells and width
        sheet name
        """
        for cell, width in cell_width:
            sheet.column_dimensions[cell].width = width

    def bold_cell(self, cells_to_bold, sheet) -> None:
        """
        bold the given cells in given sheet
        Parameters
        ----------
        list for cells to bold
        sheet name
        """
        for cell in cells_to_bold:
            sheet[cell].font = Font(bold=True, name=DEFAULT_FONT.name)

    def colour_cell(self, cells_to_colour, sheet, fill) -> None:
        """
        colour the given cells in given sheet
        Parameters
        ----------
        list for cells to color
        sheet name
        color to fill
        """
        for cell in cells_to_colour:
            sheet[cell].fill = fill

    def all_border(self, row_ranges, sheet) -> None:
        """
        create all borders for given cells in given sheet
        Parameters
        ----------
        list for row ranges
        sheet name
        """
        for row in row_ranges:
            for cells in sheet[row]:
                for cell in cells:
                    cell.border = THIN_BORDER

    def lower_border(self, cells_lower_border, sheet) -> None:
        """
        create lower cell border for given cells in given sheet
        Parameters
        ----------
        list of cells for lower border
        sheet name
        """
        for cell in cells_lower_border:
            sheet[cell].border = LOWER_BORDER

    def write_soc(self) -> None:
        """
        Write soc sheet
        """
        self.patient_info = self.read_html_tables(0)
        # write titles
        self.soc.cell(1, 1).value = "Patient Details (Epic demographics)"
        self.soc.cell(1, 3).value = "Previous testing"
        self.soc.cell(2, 1).value = "NAME"
        self.soc.cell(2, 3).value = "Alteration"
        self.soc.cell(2, 4).value = "Assay"
        self.soc.cell(2, 5).value = "Result"
        self.soc.cell(2, 6).value = "WGS concordance"
        self.soc.cell(3, 1).value = "Sex, Age, DOB"
        self.soc.cell(4, 1).value = "Phone number"
        self.soc.cell(5, 1).value = "MRN"
        self.soc.cell(6, 1).value = "NHS Number"
        self.soc.cell(8, 1).value = "Histology"
        self.soc.cell(12, 1).value = "Comments"

        # merge columns that have longer text
        self.soc.merge_cells(
            start_row=1, end_row=1, start_column=3, end_column=6
        )
        # align cells
        cell_to_align = ["C1", "C2", "D2", "E2", "F2"]
        for cell in cell_to_align:
            self.soc[cell].alignment = Alignment(
                wrapText=True, horizontal="center"
            )
            if cell != "C1":
                self.soc[cell].font = Font(italic=True)

        # titles to set to bold
        to_bold = ["A1", "A8", "A12", "A16", "C1"]
        self.bold_cell(to_bold, self.soc)

        # set column widths for readability
        cell_col_width = (
            ("A", 32),
            ("C", 16),
            ("E", 16),
            ("D", 26),
            ("F", 26),
        )
        self.set_col_width(cell_col_width, self.soc)

        # colour cells
        greenFill = PatternFill(patternType="solid", start_color="90EE90")
        colour_cells = ["C3", "D3", "E3", "F3", "C4", "D4", "E4", "F4"]
        self.colour_cell(colour_cells, self.soc, greenFill)

        # set borders around table areas
        row_ranges = []
        for i in range(1, 9):
            row_ranges.append(f"C{i}:F{i}")
        self.all_border(row_ranges, self.soc)
        cells_lower_border = ["A1", "A8", "A12"]
        self.lower_border(cells_lower_border, self.soc)

        # add dropdowns
        assay_options = (
            '"FISH,IHC,NGS,Sanger,NGS multi-gene panel,'
            "RNA fusion panel,SNP array, Methylation array,"
            "MALDI-TOF, MLPA, MS-MLPA, Chromosome breakage,"
            'Digital droplet PCR, RT-PCR, LR-PCR"'
        )
        concordance_options = (
            '"Novel,Concordant (detected),'
            "Concordant (undetected),"
            "Disconcordant (detected),"
            'Disconcordant (undetected),N/A"'
        )
        result_options = '"Detected, Not detected"'
        dropdown_pair = (
            ("D", assay_options),
            ("E", result_options),
            ("F", concordance_options),
        )
        for col, dropdown_item in dropdown_pair:
            cells_for_dropdown = []
            for i in range(3, 9):
                cells_for_dropdown.append(f"{col}{i}")

            self.get_dropdown(
                dropdown_options=dropdown_item,
                prompt="Select from the list",
                title="",
                sheet=self.soc,
                cells=cells_for_dropdown,
            )

    def write_pid_table(self, sheet_name) -> None:
        """
        write pid table in a given sheet
        Parameters
        ---------
        str for sheet name to write pid table
        """
        pid_keys = (
            (1, 1, "=SOC!A2"),
            (2, 1, "=SOC!A3"),
            (1, 3, "=SOC!A5"),
            (2, 3, "=SOC!A6"),
            (1, 5, "=SOC!A9"),
        )
        for row, col, key in pid_keys:
            sheet_name.cell(row, col).value = key

    def write_QC(self) -> None:
        """
        write QC sheet
        """
        # get QC info from html tables
        tumor_info = self.read_html_tables(1)
        sample_info = self.read_html_tables(2)
        germline_info = self.read_html_tables(3)
        seq_info = self.read_html_tables(4)
        tmb_value = self.get_tmb()
        # PID table
        self.write_pid_table(self.QC)
        self.QC.cell(15, 1).value = "QC alerts"
        self.QC.cell(16, 1).value = "None"

        # table 1
        table1_keys = (
            (1, "Diagnosis Date"),
            (2, "Tumour Received"),
            (3, "Tumour ID"),
            (4, "Presentation"),
            (5, "Diagnosis"),
            (6, "Tumour Site"),
            (7, "Tumour Type"),
            (8, "Germline Sample"),
        )
        for cell, key in table1_keys:
            self.QC.cell(4, cell).value = key

        table1_values = (
            (1, tumor_info[0]["Tumour Diagnosis Date"]),
            (2, sample_info[0]["Clinical Sample Date Time"]),
            (3, tumor_info[0]["Histopathology or SIHMDS LAB ID"]),
            (
                4,
                tumor_info[0]["Presentation"].split("_")[0]
                + " ("
                + tumor_info[0]["Primary or Metastatic"]
                + ")",
            ),
            (5, self.patient_info[0]["Clinical Indication"]),
            (6, tumor_info[0]["Tumour Topography"]),
            (
                7,
                sample_info[0]["Storage Medium"]
                + " "
                + sample_info[0]["Source"],
            ),
            (
                8,
                germline_info[0]["Storage Medium"]
                + " ("
                + germline_info[0]["Source"]
                + ")",
            ),
        )
        for cell, value in table1_values:
            self.QC.cell(5, cell).value = value

        # table 2
        table2_keys = (
            (1, "Purity (Histo)"),
            (2, "Purity (Calculated)"),
            (3, "Ploidy"),
            (4, "Total SNVs"),
            (5, "Total Indels"),
            (6, "Total SVs"),
            (7, "TMB"),
        )
        for cell, key in table2_keys:
            self.QC.cell(7, cell).value = key

        table2_values = (
            (1, sample_info[0]["Tumour Content"]),
            (2, sample_info[0]["Calculated Tumour Content"]),
            (3, sample_info[0]["Calculated Overall Ploidy"]),
            (4, seq_info[1]["Total somatic SNVs"]),
            (5, seq_info[1]["Total somatic indels"]),
            (6, seq_info[1]["Total somatic SVs"]),
            (7, str(tmb_value).strip()),
        )
        for cell, value in table2_values:
            self.QC.cell(8, cell).value = value

        # table 3
        table3_keys = (
            (1, "Sample type"),
            (2, "Mean depth, x"),
            (3, "Mapped reads, %"),
            (4, "Chimeric DNA frag, %"),
            (5, "Insert size, bp"),
            (6, "Unevenness, x"),
        )
        for cell, key in table3_keys:
            self.QC.cell(10, cell).value = key

        seq_info_title = (
            (1, "Sample type"),
            (2, "Genome-wide coverage mean, x"),
            (3, "Mapped reads, %"),
            (4, "Chimeric DNA fragments, %"),
            (5, "Insert size median, bp"),
            (6, "Unevenness of local genome coverage, x"),
        )
        for cell, title in seq_info_title:
            self.QC.cell(11, cell).value = seq_info[0][title]
            self.QC.cell(12, cell).value = seq_info[1][title]

        # titles to set to bold
        to_bold = [
            "A1",
            "A4",
            "A7",
            "A10",
            "A15",
            "B4",
            "B7",
            "B10",
            "C4",
            "C7",
            "C10",
            "D4",
            "D7",
            "D10",
            "E4",
            "E7",
            "E10",
            "F4",
            "F7",
            "F10",
            "G4",
            "G7",
            "H4",
        ]
        self.bold_cell(to_bold, self.QC)

        # set column widths for readability
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            self.QC.column_dimensions[col].width = 22

        # set borders around table areas
        row_ranges = [
            "A4:H4",
            "A5:H5",
            "A7:G7",
            "A8:G8",
            "A10:F10",
            "A11:F11",
            "A12:F12",
        ]
        self.all_border(row_ranges, self.QC)
        self.lower_border(["A15"], self.QC)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="ADD8E6")
        blue_colour_cells = [
            "A4",
            "B4",
            "C4",
            "D4",
            "E4",
            "F4",
            "G4",
            "H4",
            "A7",
            "B7",
            "C7",
            "D7",
            "E7",
            "F7",
            "G7",
            "A10",
            "B10",
            "C10",
            "D10",
            "E10",
            "F10",
        ]
        self.colour_cell(blue_colour_cells, self.QC, blueFill)

        # add dropdowns
        cells_for_QC = ["A16"]
        QC_options = (
            '"None,<30% tumour purity,SNVs low VAF (<6%),TINC (<5%),'
            'Somatic CNV, Germline CNV"'
        )
        self.get_dropdown(
            dropdown_options=QC_options,
            prompt="Select from the list",
            title="QC alerts",
            sheet=self.QC,
            cells=cells_for_QC,
        )
        # insert img from html
        self.insert_img(self.QC, "figure_9.jpg", "C16", 400, 600)
        self.insert_img(self.QC, "figure_11.jpg", "F16", 400, 600)

    def write_plot(self) -> None:
        """
        write plot sheet
        """
        # pid table
        self.write_pid_table(self.plot)
        self.plot.cell(38, 1).value = "Pertinent chromosomal CNVs"
        self.plot.cell(39, 1).value = "None"

        # titles to set to bold
        to_bold = ["A1", "A38"]
        self.bold_cell(to_bold, self.plot)

        # set column widths for readability
        cell_col_width = (
            ("A", 18),
            ("B", 22),
            ("C", 18),
            ("D", 22),
            ("E", 22),
        )
        self.set_col_width(cell_col_width, self.plot)
        # set lower border
        self.lower_border(["A38"], self.plot)

        # insert img from html file
        self.insert_img(self.plot, "figure_3.jpg", "A4", 600, 1000)
        self.insert_img(self.plot, "cropped_figure_2.jpg", "H4", 500, 500)

    def write_signatures(self) -> None:
        """
        write signatures sheet
        """
        # pid table
        self.write_pid_table(self.signatures)
        self.signatures.cell(35, 1).value = "Signature version"
        self.signatures.cell(36, 1).value = "v2 (March 2015)"
        self.signatures.cell(35, 3).value = "Pertinent signatures"
        self.signatures.cell(36, 3).value = "None"

        # titles to set to bold
        to_bold = ["A1", "A35", "C35"]
        self.bold_cell(to_bold, self.signatures)
        # set lower border
        self.lower_border(["A35", "C35"], self.signatures)

        # set column widths for readability
        cell_col_width = (
            ("A", 18),
            ("B", 22),
            ("C", 18),
            ("D", 22),
            ("E", 22),
        )
        self.set_col_width(cell_col_width, self.signatures)

        # insert img from html file
        self.insert_img(self.signatures, "figure_6.jpg", "A4", 600, 800)
        self.insert_img(self.signatures, "figure_7.jpg", "E4", 600, 800)
        self.insert_img(self.signatures, "figure_8.jpg", "M4", 600, 1100)

    def get_clnsigconf(self, clinvarID) -> str:
        """
        get the clnsigconf from clinvar file for given
        clinvar ID

        Parameters
        ----------
        int for clinvar ID
        str for clinvar ref file

        Returns
        -------
        list for CLNSIG, CLNSIGCONF
        """
        clinvar_dx = []
        for c in ["CLNSIG", "CLNSIGCONF"]:
            cmd = f"zcat {self.args.clinvar} | awk '$3=={clinvarID} \
                  {{print($8)}}' |  \
                  grep -o -P '(?<={c}=).*?(?=;)'"
            ps = subprocess.Popen(
                cmd,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
            )
            output = ps.communicate()[0]
            clinvar_dx.append(output.decode("utf-8").strip())
        if all("" == s for s in clinvar_dx):
            clinvar_dx = ""
        elif "" in clinvar_dx:
            clinvar_dx = next(s for s in clinvar_dx if s)
        else:
            clinvar_dx = clinvar_dx[1]
        return clinvar_dx

    def write_germline(self) -> None:
        """
        write germline sheet
        """
        self.write_pid_table(self.germline)
        # Germline SNV table
        snv_table_keys = (
            (1, "Gene"),
            (2, "GRCh38 Coordinates"),
            (3, "Variant"),
            (4, "Consequence"),
            (5, "Genotype"),
            (6, "Variant Class"),
            (7, "Actionability"),
            (8, "Role in Cancer"),
            (9, "ClinVar"),
            (10, "gnomAD"),
            (11, "Tumour VAF"),
        )
        for cell, key in snv_table_keys:
            self.germline.cell(4, cell).value = key

        # populate germline table
        germline_table = pd.read_csv(self.args.variant, sep=",")
        germline_table = germline_table[germline_table["Origin"] == "germline"]
        germline_table.reset_index(drop=True, inplace=True)

        # get the clnsigconf from clinvar file based on clinvar ID
        clinvarID = list(germline_table["ClinVar ID"])
        d = []
        for cid in clinvarID:
            d.append(
                {
                    "ClinVar ID": cid,
                    "clnsigconf": self.get_clnsigconf(cid),
                }
            )
        clinvar_df = pd.DataFrame(d)
        germline_table = germline_table.merge(
            clinvar_df, on="ClinVar ID", how="left"
        )

        # split the col to get gnomAD
        germline_table[["GE", "gnomAD"]] = germline_table[
            "Population germline allele frequency (GE | gnomAD)"
        ].str.split("|", expand=True)
        germline_table.drop(
            ["GE", "Population germline allele frequency (GE | gnomAD)"],
            axis=1,
            inplace=True,
        )
        germline_table.loc[:, "Variant Class"] = ""
        germline_table.loc[:, "Actionability"] = ""
        germline_table = germline_table[
            [
                "Gene",
                "GRCh38 coordinates;ref/alt allele",
                "CDS change and protein change",
                "Predicted consequences",
                "Genotype",
                "Variant Class",
                "Actionability",
                "Gene mode of action",
                "clnsigconf",
                "gnomAD",
            ]
        ]

        # write df into excel sheet
        num_gene = germline_table.shape[0]
        rows = dataframe_to_rows(germline_table)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if c_idx != 1 and r_idx != 1:
                    self.germline.cell(
                        row=r_idx + 2, column=c_idx - 1, value=value
                    )

        self.germline.cell(
            num_gene + 6, 1
        ).value = "Pertinent variants/feedback"
        self.germline.cell(num_gene + 7, 1).value = "None"

        # titles to set to bold
        to_bold = [
            "A1",
            "A4",
            "B4",
            "A10",
            "C4",
            "D4",
            "E4",
            "F4",
            "G4",
            "H4",
            "I4",
            "J4",
            "K4",
            f"C{num_gene+6}",
        ]
        self.bold_cell(to_bold, self.germline)
        # set border
        cells_lower_border = [
            f"A{num_gene+6}",
        ]
        self.lower_border(cells_lower_border, self.germline)

        # set borders around table areas
        row_ranges = []
        for i in range(4, num_gene + 5):
            row_ranges.append(f"A{i}:K{i}")
        self.all_border(row_ranges, self.germline)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="ADD8E6")
        blue_colour_cells = []
        for i in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            blue_colour_cells.append(f"{i}4")
        self.colour_cell(blue_colour_cells, self.germline, blueFill)

        # set column widths for readability
        cell_col_width = (
            ("A", 20),
            ("B", 18),
            ("C", 22),
            ("D", 14),
            ("F", 22),
            ("G", 18),
            ("H", 12),
            ("I", 22),
            ("K", 12),
        )
        self.set_col_width(cell_col_width, self.germline)
        # dropdowns
        cells_for_variant_class = []
        for i in range(5, num_gene + 5):
            cells_for_variant_class.append(f"F{i}")
        self.variant_class_options = (
            '"Pathogenic", "Likely pathogenic",'
            '"Uncertain", "Likely passenger",'
            '"Likely artefact"'
        )
        self.get_dropdown(
            dropdown_options=self.variant_class_options,
            prompt="Select from the list",
            title="Variant class",
            sheet=self.germline,
            cells=cells_for_variant_class,
        )
        cells_for_action = []
        for i in range(5, num_gene + 5):
            cells_for_action.append(f"G{i}")
        self.action_options = (
            '"1. Predicts therapeutic response,'
            " 2. Prognostic, 3. Defines diagnosis group"
            ', 4. Eligibility for trial, 5. Other"'
        )
        self.get_dropdown(
            dropdown_options=self.action_options,
            prompt="Select from the list",
            title="Actionability",
            sheet=self.germline,
            cells=cells_for_action,
        )

    def write_summary(self) -> None:
        """
        Write summary sheet
        """
        # pid table
        self.write_pid_table(self.summary)
        # germline table
        self.summary.cell(23, 1).value = "Germline SNV"
        snv_table_keys = (
            (1, "Gene"),
            (2, "GRCh38 Coordinates"),
            (3, "Variant"),
            (4, "Consequence"),
            (5, "Zygosity"),
            (6, "Variant Class"),
            (7, "Actionability"),
            (8, "Comments"),
        )
        for cell, key in snv_table_keys:
            self.summary.cell(24, cell).value = key
        # snv table
        self.summary.cell(31, 1).value = "Somatic SNV"
        snv_table_keys = (
            (1, "Gene"),
            (2, "GRCh38 Coordinates"),
            (3, "Mutation"),
            (4, "Consequence"),
            (5, "VAF"),
            (6, "Variant Class"),
            (7, "Actionability"),
            (8, "Comments"),
        )
        for cell, key in snv_table_keys:
            self.summary.cell(32, cell).value = key

        # cnv sv table
        self.summary.cell(44, 1).value = "Somatic CNV_SV"
        cnv_sv_table_key = (
            (1, "Gene/Locus"),
            (2, "GRCh38 Coordinates"),
            (3, "Cytological Bands"),
            (4, "Variant Type"),
            (5, "Consequence"),
            (6, "Variant Class"),
            (7, "Actionability"),
            (8, "Comments"),
        )
        for cell, key in cnv_sv_table_key:
            self.summary.cell(45, cell).value = key

        # titles to set to bold
        to_bold = [
            "A1",
            "A23",
            "A24",
            "A31",
            "A32",
            "A44",
            "A45",
            "B24",
            "B32",
            "B45",
            "C24",
            "C32",
            "C45",
            "D24",
            "D32",
            "D45",
            "E24",
            "E32",
            "E45",
            "F24",
            "F32",
            "F45",
            "G24",
            "G32",
            "G45",
            "H24",
            "H32",
            "H45",
            "A58",
            "A67",
            "A78",
            "A89",
        ]
        self.bold_cell(to_bold, self.summary)

        # set column widths for readability
        cell_col_width = (
            ("A", 26),
            ("B", 20),
            ("C", 22),
            ("D", 24),
            ("F", 24),
            ("G", 24),
            ("H", 24),
        )
        self.set_col_width(cell_col_width, self.summary)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="ADD8E6")
        colour_cells = []
        for cell in [24, 32, 45]:
            for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                colour_cells.append(f"{col}{cell}")
        self.colour_cell(colour_cells, self.summary, blueFill)

        # set borders around table areas
        row_ranges = []
        for cell in range(24, 30):
            row_ranges.append(f"A{cell}:H{cell}")
        for cell in range(32, 43):
            row_ranges.append(f"A{cell}:H{cell}")
        for cell in range(45, 57):
            row_ranges.append(f"A{cell}:H{cell}")
        self.all_border(row_ranges, self.summary)
        # insert img from html file
        self.insert_img(self.summary, "figure_3.jpg", "A4", 300, 700)
        self.insert_img(self.summary, "cropped_figure_2.jpg", "F4", 350, 350)

        # for scientist to copy
        table_title = (
            (58, "Germline_SNV"),
            (67, "Somatic_SNV"),
            (78, "CNV"),
            (89, "SV"),
        )
        for cell, title in table_title:
            self.summary.cell(cell, 1).value = title
        germline_snv_keys = (
            (1, "Gene"),
            (2, "GRCh38 Coordinates"),
            (3, "Variant"),
            (4, "Consequence"),
            (5, "Genotype"),
            (6, "Variant Class"),
            (7, "Actionability"),
            (8, "Role in Cancer"),
            (9, "ClinVar"),
            (10, "gnomAD"),
            (11, "Tumour VAF"),
        )
        for cell, key in germline_snv_keys:
            self.summary.cell(59, cell).value = key

        sv_titles = (
            (68, list(self.df_SNV_copy.columns)),
            (79, list(self.df_gain_copy.columns)),
            (90, list(self.df_SV_summary.columns)),
        )

        for cell, title in sv_titles:
            for idx, key in enumerate(title):
                self.summary.cell(cell, idx + 1).value = key

        # add formula
        for i in range(25, 30):
            for idx, col in enumerate(["A", "B", "C", "D", "E", "F", "G"]):
                self.summary.cell(i, idx + 1).value = f"={col}{i+35}"
        snv_pair = (
            (1, "B"),
            (2, "C"),
            (3, "D"),
            (4, "E"),
            (5, "F"),
            (6, self.class_col),
            (7, self.action_col),
            (8, self.comment_col),
        )
        for i in range(33, 43):
            for pair1, pair2 in snv_pair:
                self.summary.cell(i, pair1).value = f"={pair2}{i+36}"

        cnv_pair = (
            (1, "C"),
            (2, "D"),
            (3, "E"),
            (4, "F"),
            (6, "J"),
            (7, "K"),
            (8, "L"),
        )
        fill1 = PatternFill(patternType="solid", start_color="E7B8C8")
        for i in range(46, 52):
            for pair1, pair2 in cnv_pair:
                self.summary.cell(i, pair1).value = f"={pair2}{i+34}"
                self.summary.cell(i, pair1).fill = fill1
        cell_to_colour = ["E46", "E47", "E48", "E49", "E50", "E51"]
        self.colour_cell(cell_to_colour, self.summary, fill1)
        if self.fusion_count == 0:
            sv_pair = (
                (1, "C"),
                (2, "D"),
                (3, "E"),
                (4, "F"),
                (6, self.col_letter_class),
                (7, self.col_letter_action),
                (8, self.col_letter_comment),
            )
        else:
            sv_pair = (
                (1, "C"),
                (2, "D"),
                (3, "E"),
                (4, "F"),
                (5, "G"),
                (6, self.col_letter_class),
                (7, self.col_letter_action),
                (8, self.col_letter_comment),
            )
        fill2 = PatternFill(patternType="solid", start_color="C8F7E3")
        for i in range(52, 57):
            for pair1, pair2 in sv_pair:
                self.summary.cell(i, pair1).value = f"={pair2}{i+39}"
                self.summary.cell(i, pair1).fill = fill2

    def aggregate_variants(self, col, df) -> pd.DataFrame:
        """
        function to aggregate all variants per gene
        Parameters
        ----------
        list of col to aggregate
        pd.DataFrame to aggregate

        Returns
        -------
        pd.DataFrame - aggregated df
        """
        n = 0
        for c in col:
            temp_df = df.groupby("Gene", as_index=False).agg(
                {c: lambda x: ",".join(x)}
            )
            if n == 0:
                new_df = temp_df
            else:
                new_df = pd.merge(new_df, temp_df, on="Gene", how="left")
            n = n + 1
        return new_df

    def concat_dataframe(self, col) -> pd.DataFrame:
        """
        function to concat df for each gene in SV
        Parameters
        ----------
        list of col to concat

        Returns
        -------
        concat df
        """
        appended_data = []
        for g in self.gene_col:
            selected_col = [g] + col
            temp_df = self.df_SV_copy[selected_col]
            temp_df.rename(columns={temp_df.columns[0]: "Gene"}, inplace=True)
            appended_data.append(temp_df)
        concat_df = pd.concat(appended_data)
        concat_df = concat_df.reset_index(drop=True)
        return concat_df

    def write_refgene(self) -> None:
        """
        write RefGene sheet
        """
        ref_sheets = (
            ("cosmic", "COSMIC"),
            ("paed", "Paediatric"),
            ("sarc", "Sarcoma"),
            ("neuro", "Neuro"),
            ("ovarian", "Ovarian"),
            ("haem", "Haem-onc"),
        )
        # write sheets
        for ref, tab in ref_sheets:
            df = pd.read_excel(self.args.refgene_group, sheet_name=ref)
            if "RefGene Group" in df.columns:
                df = df.drop("RefGene Group", axis=1)
            joined_SNV = self.df_SNV_copy[
                [
                    "Gene",
                    "GRCh38 coordinates",
                    "Variant",
                    "Predicted consequences",
                    "VAF",
                ]
            ]
            joined_SNV["VAF"] = joined_SNV["VAF"].astype("str")
            constant_col1 = [
                "GRCh38 coordinates",
                "Variant",
                "Predicted consequences",
                "VAF",
            ]
            # concat variants per gene
            final_SNV = self.aggregate_variants(constant_col1, joined_SNV)
            merged_df1 = pd.merge(df, final_SNV, on="Gene", how="left")

            CNV = pd.concat([self.df_gain_copy, self.df_loss_copy])
            CNV = CNV[["Gene", "Chromosomal bands", "Type", "Copy Number"]]
            CNV["Copy Number"] = CNV["Copy Number"].astype("str")
            constant_col2 = ["Chromosomal bands", "Type", "Copy Number"]
            final_CNV = self.aggregate_variants(constant_col2, CNV)
            merged_df2 = pd.merge(merged_df1, final_CNV, on="Gene", how="left")
            if self.max_num_gene == 1:
                if self.fusion_count == 1:
                    joined_SV = self.df_SV_copy[
                        ["Gene", "GRCh38 coordinates", "Type", "Fusion"]
                    ]
                    joined_SV["Fusion"] = joined_SV["Fusion"].fillna("null")
                    constant_col3 = ["GRCh38 coordinates", "Type", "Fusion"]
                    final_SV = self.aggregate_variants(
                        constant_col3, joined_SV
                    )
                else:
                    joined_SV = self.df_SV_copy[
                        ["Gene", "GRCh38 coordinates"] + self.fusion_col
                    ]
                    for f in self.fusion_col:
                        joined_SV[f] = joined_SV[f].fillna("null")
                    constant_col3 = ["GRCh38 coordinates"] + self.fusion_col
                    final_SV = self.aggregate_variants(
                        constant_col3, joined_SV
                    )

            else:
                if self.fusion_count == 1:
                    constant_col3 = ["GRCh38 coordinates", "Type", "Fusion"]
                    joined_SV = self.concat_dataframe(constant_col3)
                    joined_SV["Fusion"] = joined_SV["Fusion"].fillna("null")
                    final_SV = self.aggregate_variants(
                        constant_col3, joined_SV
                    )
                else:
                    constant_col3 = ["GRCh38 coordinates"] + self.fusion_col
                    joined_SV = self.concat_dataframe(constant_col3)
                    for f in self.fusion_col:
                        joined_SV[f] = joined_SV[f].fillna("null")
                    final_SV = self.aggregate_variants(
                        constant_col3, joined_SV
                    )
            merged_df3 = pd.merge(merged_df2, final_SV, on="Gene", how="left")
            merged_df3.rename(
                columns={
                    "GRCh38 coordinates_x": "GRCh38 coordinates_SNV",
                    "GRCh38 coordinates_y": "GRCh38 coordinates_SV",
                    "Type_x": "Type_CNV",
                    "Type_y": "Type_SV",
                },
                inplace=True,
            )
            merged_df3.to_excel(self.writer, sheet_name=tab, index=False)
            # format sheet
            ref_sheet = self.writer.sheets[tab]
            ref_sheet.sheet_properties.tabColor = "FF0000"
            max_col = merged_df3.shape[1]
            max_col_letter = get_column_letter(max_col)
            filters = ref_sheet.auto_filter
            filters.ref = f"A:{max_col_letter}"
            if ref == "paed":
                col_color = (
                    (
                        "G",
                        "J",
                        PatternFill(patternType="solid", start_color="FFDBBB"),
                    ),
                    (
                        "K",
                        "M",
                        PatternFill(patternType="solid", start_color="c4d9ef"),
                    ),
                    (
                        "N",
                        max_col_letter,
                        PatternFill(patternType="solid", start_color="9FE2BF"),
                    ),
                )
                for start_col, end_col, fill_color in col_color:
                    self.color_col(
                        ref_sheet,
                        start_col,
                        end_col,
                        merged_df3.shape[0] + 1,
                        fill_color,
                    )
            elif ref == "cosmic":
                col_color = (
                    (
                        "J",
                        "M",
                        PatternFill(patternType="solid", start_color="FFDBBB"),
                    ),
                    (
                        "N",
                        "P",
                        PatternFill(patternType="solid", start_color="c4d9ef"),
                    ),
                    (
                        "Q",
                        max_col_letter,
                        PatternFill(patternType="solid", start_color="9FE2BF"),
                    ),
                )
                for start_col, end_col, fill_color in col_color:
                    self.color_col(
                        ref_sheet,
                        start_col,
                        end_col,
                        merged_df3.shape[0] + 1,
                        fill_color,
                    )
            else:
                col_color = (
                    (
                        "I",
                        "L",
                        PatternFill(patternType="solid", start_color="FFDBBB"),
                    ),
                    (
                        "M",
                        "O",
                        PatternFill(patternType="solid", start_color="c4d9ef"),
                    ),
                    (
                        "P",
                        max_col_letter,
                        PatternFill(patternType="solid", start_color="9FE2BF"),
                    ),
                )
                for start_col, end_col, fill_color in col_color:
                    self.color_col(
                        ref_sheet,
                        start_col,
                        end_col,
                        merged_df3.shape[0] + 1,
                        fill_color,
                    )

    def lookup(
        self, df_to_check, ref_df, col_to_map, ref_col, lookup_col
    ) -> list:
        """
        get the list of look up col

        Parameters
        ----------
        pd.DataFrame to check
        pd.DataFrame to refer
        str col name to map
        str col name as ref col
        str col to look up

        Returns
        -------
        list - result of look up
        """
        return df_to_check[col_to_map].map(
            ref_df.set_index(ref_col)[lookup_col]
        )

    def get_col_letter(self, worksheet, col_name) -> str:
        """
        Getting the column letter with specific col name

        Parameters
        ----------
        worksheet: openpyxl.Writer
               writer object of current sheet
        col_name: str
               name of column to get col letter
        Return
        -------
        str
            column letter for specific column name
        """
        col_letter = None
        for column_cell in worksheet.iter_cols(1, worksheet.max_column):
            if column_cell[0].value == col_name:
                col_letter = column_cell[0].column_letter

        return col_letter

    def write_SNV(self) -> None:
        """
        write SNV sheet
        """
        df_cosmic = pd.read_excel(self.args.refgene_group, sheet_name="cosmic")
        df_paed = pd.read_excel(self.args.refgene_group, sheet_name="paed")
        df_sarc = pd.read_excel(self.args.refgene_group, sheet_name="sarc")
        df_neuro = pd.read_excel(self.args.refgene_group, sheet_name="neuro")
        df_ovarian = pd.read_excel(
            self.args.refgene_group, sheet_name="ovarian"
        )
        df_haem = pd.read_excel(self.args.refgene_group, sheet_name="haem")
        for df in [
            df_cosmic,
            df_paed,
            df_sarc,
            df_neuro,
            df_ovarian,
            df_haem,
        ]:
            if "Entities" in list(df.columns):
                df["Entities"] = df["Entities"].fillna("*")
            if "Driver" in list(df.columns):
                df["Driver"] = df["Driver"].fillna("*")
        df_hotspots = pd.read_csv(self.args.hotspots)
        df = pd.read_csv(self.args.variant, sep=",")
        # select only somatic rows
        df = df[df["Origin"] == "somatic"]
        df.reset_index(drop=True, inplace=True)
        num_variant = df.shape[0]
        df[["c_dot", "p_dot"]] = df["CDS change and protein change"].str.split(
            r"(?=p)", n=1, expand=True
        )
        df["c_dot"] = df["c_dot"].str.replace("(;$)", "", regex=True)

        # look up genes from df_refgene
        self.lookup_refgene = (
            ("COSMIC", df_cosmic, "Entities"),
            ("Paed", df_paed, "Driver"),
            ("Sarc", df_sarc, "Driver"),
            ("Neuro", df_neuro, "Driver"),
            ("Ovary", df_ovarian, "Driver"),
            ("Haem", df_haem, "Driver"),
        )
        for j, k, v in self.lookup_refgene:
            df[j] = self.lookup(df, k, "Gene", "Gene", v)
            df[j] = df[j].fillna("-")

        df = df.replace([None], [""], regex=True)
        df["MTBP c."] = df["Gene"] + ":" + df["c_dot"]
        df["MTBP p."] = df["Gene"] + ":" + df["p_dot"]
        df[["HS p.", "col1", "col2"]] = df["MTBP p."].str.split(
            r"([^\d]+)$", expand=True
        )
        # look up from hotspots
        lookup_dict_hotspots = {
            "HS_Sample": "HS_Samples",
            "HS_Tumour": "HS_Tumor Type Composition",
        }
        for k, v in lookup_dict_hotspots.items():
            df[k] = self.lookup(df, df_hotspots, "HS p.", "HS_PROTEIN_ID", v)
            df[k] = df[k].fillna("-")
        error = []
        df["con_count"] = df["Predicted consequences"].str.count(r"\;")
        if df["con_count"].max() > 0:
            df[["Predicted consequences", "Error flag"]] = df[
                "Predicted consequences"
            ].str.split(";", expand=True)
            error = ["Error flag"]
        loh = []
        df["VAF"] = df["VAF"].astype("str")
        df["VAF_count"] = df["VAF"].str.count(r"\;")
        if df["VAF_count"].max() > 0:
            df[["VAF", "LOH"]] = df["VAF"].str.split(";", expand=True)
            loh = ["LOH"]
        df.loc[:, "Variant class"] = ""
        df.loc[:, "Actionability"] = ""
        df.loc[:, "Comments"] = ""
        df = df[
            [
                "Domain",
                "Gene",
                "GRCh38 coordinates;ref/alt allele",
                "CDS change and protein change",
                "Predicted consequences",
                "VAF",
            ]
            + loh
            + error
            + [
                "Alt allele/total read depth",
                "Gene mode of action",
                "Variant class",
                "Actionability",
                "Comments",
                "COSMIC",
                "Paed",
                "Sarc",
                "Neuro",
                "Ovary",
                "Haem",
                "HS_Sample",
                "HS_Tumour",
                "MTBP c.",
                "MTBP p.",
            ]
        ]
        df.rename(
            columns={
                "GRCh38 coordinates;ref/alt allele": "GRCh38 coordinates",
                "CDS change and protein change": "Variant",
            },
            inplace=True,
        )
        df.sort_values(
            ["Domain", "VAF"], ascending=[True, False], inplace=True
        )
        df["VAF"] = df["VAF"].astype(float)
        self.df_SNV_copy = df
        df.to_excel(self.writer, sheet_name="SNV", index=False)
        self.SNV = self.writer.sheets["SNV"]
        cell_col_width = (
            ("B", 12),
            ("C", 28),
            ("D", 28),
            ("E", 18),
            ("F", 14),
            ("J", 20),
            ("K", 20),
            ("L", 20),
            ("M", 20),
            ("N", 20),
            ("O", 14),
            ("P", 22),
            ("Q", 26),
            ("R", 18),
            ("S", 18),
            ("T", 16),
            ("U", 16),
            ("V", 18),
            ("W", 18),
        )
        self.set_col_width(cell_col_width, self.SNV)

        # get max col for dropdown
        max_col = df.shape[1]
        max_col_letter = get_column_letter(max_col)
        # add filter in col
        filters = self.SNV.auto_filter
        filters.ref = f"A:{max_col_letter}"

        # add dropdowns
        self.add_dropdonws_sheet(self.SNV, num_variant)
        self.class_col = self.get_col_letter(self.SNV, "Variant class")
        self.action_col = self.get_col_letter(self.SNV, "Actionability")
        self.comment_col = self.get_col_letter(self.SNV, "Comments")
        cosmic_col = self.get_col_letter(self.SNV, "COSMIC")
        haem_col = self.get_col_letter(self.SNV, "Haem")
        sample_col = self.get_col_letter(self.SNV, "HS_Sample")
        tumor_col = self.get_col_letter(self.SNV, "HS_Tumour")
        cdot_col = self.get_col_letter(self.SNV, "MTBP c.")
        pdot_col = self.get_col_letter(self.SNV, "MTBP p.")

        col_color = (
            (
                self.class_col,
                self.comment_col,
                PatternFill(patternType="solid", start_color="FFDBBB"),
            ),
            (
                cosmic_col,
                haem_col,
                PatternFill(patternType="solid", start_color="c4d9ef"),
            ),
            (
                sample_col,
                tumor_col,
                PatternFill(patternType="solid", start_color="00FFFF"),
            ),
            (
                cdot_col,
                pdot_col,
                PatternFill(patternType="solid", start_color="dabcff"),
            ),
        )
        for start_col, end_col, fill_color in col_color:
            self.color_col(
                self.SNV, start_col, end_col, num_variant + 1, fill_color
            )
        self.SNV.freeze_panes = self.SNV["E1"]
        self.data_bar("F2", f"F{num_variant + 1}", self.SNV)

    def color_col(
        self, sheet, start_col, end_col, max_row, color_to_fill
    ) -> None:
        """
        color the cols in given sheet
        Parameters
        ----------
        sheet name
        str for start col
        str for end col
        int for max row
        PatternFill for hex color code

        """
        for row in sheet[f"{start_col}1:{end_col}{max_row}"]:
            for cell in row:
                cell.fill = color_to_fill

    def write_SV(self) -> None:
        """
        write SV sheet
        """
        df_SV = pd.read_csv(self.args.SV, sep=",")
        # subset df for SV
        df_SV = df_SV[~df_SV["Type"].str.lower().str.contains("loss|loh|gain")]
        # split fusion columns
        df_SV["fusion_count"] = df_SV["Type"].str.count(r"\;")
        self.fusion_count = df_SV["fusion_count"].max()
        if self.fusion_count == 1:
            df_SV[["Type", "Fusion"]] = df_SV.Type.str.split(";", expand=True)
        else:
            self.fusion_col = []
            for i in range(self.fusion_count):
                self.fusion_col.append(f"Fusion_{i+1}")
            self.fusion_col.insert(0, "Type")
            df_SV[self.fusion_col] = df_SV.Type.str.split(";", expand=True)
        df_SV[["Paired reads", "Split reads"]] = df_SV[
            "Confidence/support"
        ].str.split(";", expand=True)
        df_SV[["col1", "Paired reads"]] = df_SV["Paired reads"].str.split(
            "-", expand=True
        )
        df_SV[["col2", "Split reads"]] = df_SV["Split reads"].str.split(
            "-", expand=True
        )
        # get thousands separator
        df_SV["Size"] = df_SV.apply(
            lambda x: "{:,.0f}".format(x["Size"]), axis=1
        )
        # replace nan in size with -
        df_SV["Size"] = df_SV["Size"].str.replace("nan", "-")
        # get gene counts and look up for each gene
        df_SV["gene_count"] = df_SV["Gene"].str.count(r"\;")
        self.max_num_gene = df_SV["gene_count"].max() + 1
        # split gene col and create look up col for them
        if self.max_num_gene == 1:
            # look up genes from df_refgene
            for j, k, v in self.lookup_refgene:
                df_SV[j] = self.lookup(df_SV, k, "Gene", "Gene", v)
                df_SV[j] = df_SV[j].fillna("-")
        else:
            self.gene_col = []
            for i in range(self.max_num_gene):
                self.gene_col.append(f"Gene_{i+1}")
            df_SV[self.gene_col] = df_SV["Gene"].str.split(";", expand=True)
            for g in range(self.max_num_gene):
                for j, k, v in self.lookup_refgene:
                    df_SV[f"{j}_{g+1}"] = self.lookup(
                        df_SV, k, self.gene_col[g], "Gene", v
                    )
                    df_SV[f"{j}_{g+1}"] = df_SV[f"{j}_{g+1}"].fillna("-")
        df_SV.loc[:, "Variant class"] = ""
        df_SV.loc[:, "Actionability"] = ""
        df_SV.loc[:, "Comments"] = ""
        to_lookup = ("COSMIC", "Paed", "Sarc", "Neuro", "Ovary", "Haem")
        lookup_col = [col for col in df_SV if col.startswith(to_lookup)]
        self.df_SV_copy = df_SV
        if self.fusion_count == 1:
            selected_col = [
                "Event domain",
                "Impacted transcript region",
                "Gene",
                "GRCh38 coordinates",
                "Chromosomal bands",
                "Type",
                "Fusion",
                "Size",
                "Population germline allele frequency (GESG | GECG for somatic SVs or AF | AUC for germline CNVs)",
                "Paired reads",
                "Split reads",
                "Gene mode of action",
                "Variant class",
                "Actionability",
                "Comments",
            ] + lookup_col

        else:
            selected_col = (
                [
                    "Event domain",
                    "Impacted transcript region",
                    "Gene",
                    "GRCh38 coordinates",
                    "Chromosomal bands",
                ]
                + self.fusion_col
                + [
                    "Size",
                    "Population germline allele frequency (GESG | GECG for somatic SVs or AF | AUC for germline CNVs)",
                    "Paired reads",
                    "Split reads",
                    "Gene mode of action",
                    "Variant class",
                    "Actionability",
                    "Comments",
                ]
                + lookup_col
            )
        df_SV = df_SV[selected_col]
        self.df_SV_summary = df_SV
        # write each df into sheet
        num_variant = df_SV.shape[0]
        max_col = df_SV.shape[1]
        df_SV.to_excel(self.writer, sheet_name="SV", index=False)
        self.SV = self.writer.sheets["SV"]
        cell_col_width = (
            ("B", 12),
            ("C", 22),
            ("D", 22),
            ("E", 20),
            ("G", 16),
            ("H", 16),
            ("I", 18),
            ("J", 18),
            ("K", 18),
            ("L", 18),
            ("M", 18),
            ("N", 18),
            ("O", 18),
        )
        self.set_col_width(cell_col_width, self.SV)
        max_col_letter = get_column_letter(max_col)
        filters = self.SV.auto_filter
        filters.ref = f"A:{max_col_letter}"
        # add dropdowns
        self.add_dropdonws_sheet(self.SV, num_variant)
        col_letter_lookup1 = self.get_col_letter(self.SV, lookup_col[0])
        col_letter_lookup2 = self.get_col_letter(self.SV, lookup_col[-1])
        self.col_letter_comment = self.get_col_letter(self.SV, "Comments")
        self.col_letter_class = self.get_col_letter(self.SV, "Variant class")
        self.col_letter_action = self.get_col_letter(self.SV, "Actionability")

        col_color = (
            (
                self.col_letter_class,
                self.col_letter_comment,
                PatternFill(patternType="solid", start_color="FFDBBB"),
            ),
            (
                col_letter_lookup1,
                col_letter_lookup2,
                PatternFill(patternType="solid", start_color="c4d9ef"),
            ),
        )
        for start_col, end_col, fill_color in col_color:
            self.color_col(
                self.SV, start_col, end_col, num_variant + 1, fill_color
            )
        self.SV.freeze_panes = self.SV["F1"]

    def write_gain_loss(self) -> None:
        """
        write GAIN and LOSS sheets
        """
        df = pd.read_csv(self.args.SV, sep=",")
        df_loss = df[df["Type"].str.lower().str.contains("loss|loh")]
        df_loss.reset_index(drop=True, inplace=True)
        df_gain = df[df["Type"].str.lower().str.contains("gain")]
        df_gain.reset_index(drop=True, inplace=True)

        # look up genes from df_refgene
        for df in [df_loss, df_gain]:
            for j, k, v in self.lookup_refgene:
                df[j] = self.lookup(df, k, "Gene", "Gene", v)
                df[j] = df[j].fillna("-")
            df.loc[:, "Variant class"] = ""
            df.loc[:, "Actionability"] = ""
            df.loc[:, "Comments"] = ""
            df[["Type", "Copy Number"]] = df.Type.str.split(
                r"\(|\)", expand=True
            ).iloc[:, [0, 1]]
            df["Copy Number"] = df["Copy Number"].astype(int)
            df["Size"] = df.apply(
                lambda x: "{:,.0f}".format(x["Size"]), axis=1
            )
            if list(df["Type"].unique()) == ["GAIN"]:
                df.sort_values(
                    ["Event domain", "Copy Number"],
                    ascending=[True, False],
                    inplace=True,
                )
            else:
                df.sort_values(
                    ["Event domain", "Copy Number"],
                    ascending=[True, True],
                    inplace=True,
                )
        # subset df
        selected_col = [
            "Event domain",
            "Impacted transcript region",
            "Gene",
            "GRCh38 coordinates",
            "Chromosomal bands",
            "Type",
            "Copy Number",
            "Size",
            "Gene mode of action",
            "Variant class",
            "Actionability",
            "Comments",
            "COSMIC",
            "Paed",
            "Sarc",
            "Neuro",
            "Ovary",
            "Haem",
        ]
        df_loss = df_loss[selected_col]
        df_gain = df_gain[selected_col]
        df_to_write = (
            (df_loss, "LOSS"),
            (df_gain, "GAIN"),
        )
        self.df_gain_copy = df_gain
        self.df_loss_copy = df_loss
        # write each df into sheet
        for df, sheet_name in df_to_write:
            num_variant = df.shape[0]
            df.to_excel(self.writer, sheet_name=sheet_name, index=False)
            sheet = self.writer.sheets[sheet_name]
            cell_col_width = (
                ("B", 12),
                ("C", 16),
                ("D", 22),
                ("E", 20),
                ("G", 16),
                ("H", 14),
                ("I", 22),
                ("J", 20),
                ("K", 20),
                ("L", 20),
                ("M", 22),
                ("N", 20),
                ("O", 16),
                ("P", 16),
                ("Q", 16),
                ("R", 16),
            )
            self.set_col_width(cell_col_width, sheet)

            # add dropdowns
            self.add_dropdonws_sheet(sheet, num_variant)
            col_color = (
                (
                    "J",
                    "L",
                    PatternFill(patternType="solid", start_color="FFDBBB"),
                ),
                (
                    "M",
                    "R",
                    PatternFill(patternType="solid", start_color="c4d9ef"),
                ),
            )
            for start_col, end_col, fill_color in col_color:
                self.color_col(
                    sheet, start_col, end_col, num_variant + 1, fill_color
                )
            sheet.freeze_panes = sheet["F1"]
            max_col = df.shape[1]
            max_col_letter = get_column_letter(max_col)
            filters = sheet.auto_filter
            filters.ref = f"A:{max_col_letter}"
            # make left align for Copy Number col
            cell_to_align = []
            for i in range(2, num_variant + 2):
                cell_to_align.append(f"G{i}")
            for cell in cell_to_align:
                sheet[cell].alignment = Alignment(
                    wrapText=True, horizontal="left"
                )

    def get_dropdown(
        self, dropdown_options, prompt, title, sheet, cells
    ) -> None:
        """
        create the dropdowns items for designated cells

        Parameters
        ----------
        dropdown_options: str
            str containing dropdown items
        prompt: str
            prompt message for dropdown
        title: str
            title message for dropdown
        sheet: openpyxl.Writer writer object
            current worksheet
        cells: list
            list of cells to add dropdown
        """
        options = dropdown_options
        val = DataValidation(type="list", formula1=options, allow_blank=True)
        val.prompt = prompt
        val.promptTitle = title
        sheet.add_data_validation(val)
        for cell in cells:
            val.add(sheet[cell])
        val.showInputMessage = True
        val.showErrorMessage = True

    def insert_img(self, sheet, img_to_insert, cell_to_insert, h, w) -> None:
        """
        insert the img downloaded from html into spreadsheet
        """
        ws = sheet
        img = drawing.image.Image(img_to_insert)
        img.height = h
        img.width = w
        img.anchor = cell_to_insert
        ws.add_image(img)

    def add_dropdonws_sheet(self, sheet_name, num_variant) -> None:
        """
        adding variant_class and
        actionability dropdonws to selected sheets
        Parameters
        ----------
        sheet name to add dropdowns
        num of variant in that sheet
        """
        col_letter_class = self.get_col_letter(sheet_name, "Variant class")
        col_letter_action = self.get_col_letter(sheet_name, "Actionability")
        cells_for_class = []
        for i in range(2, num_variant + 2):
            cells_for_class.append(f"{col_letter_class}{i}")
        self.get_dropdown(
            dropdown_options=self.variant_class_options,
            prompt="Select from the list",
            title="Variant class",
            sheet=sheet_name,
            cells=cells_for_class,
        )
        cells_for_action = []
        for i in range(2, num_variant + 2):
            cells_for_action.append(f"{col_letter_action}{i}")

        self.get_dropdown(
            dropdown_options=self.action_options,
            prompt="Select from the list",
            title="Actionability",
            sheet=sheet_name,
            cells=cells_for_action,
        )

    def data_bar(self, start_cell, end_cell, sheet_name) -> None:
        """
        add databar in the sheet
        Parameters
        ----------
        str for start cell
        str for end cell
        str for sheet name
        """
        rule = DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="FF3361",
        )
        sheet_name.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)


def main():
    # generate output Excel file
    excel_handler = excel()
    excel_handler.generate()


if __name__ == "__main__":
    main()
