import argparse
import re
import subprocess
import urllib.request
import numpy as np
from openpyxl import drawing
from openpyxl.styles import Alignment, Border, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
from bs4 import BeautifulSoup
from PIL import Image

# ref files
HTOSPOTS_REF = "./resources/Hotspots.csv"
REFGENEGP_REF = "./resources/RefGene_Groups.csv"
CLINVAR_REF = "./resources/clinvar_20240407_GRCh37.vcf.gz"

# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)
DEFAULT_FONT.name = "Calibri"


class excel:
    """
    Functions for wrangling input csv files, ref files and html files and
    writing output xlsx file

    Attributes
    ----------
    args : argparse.Namespace
        arguments passed from command line
    writer : pandas.io.excel._openpyxl.OpenpyxlWriter
        writer object for writing Excel data to file
    workbook : openpyxl.workbook.workbook.Workbook
        openpyxl workbook object for interacting with per-sheet writing and
        formatting of output Excel file

    Outputs
    -------
    {args.output}.xlsx : file
        Excel file with variants and structural variants
    """

    def __init__(self) -> None:
        self.args = self.parse_args()
        self.writer = pd.ExcelWriter(self.args.output, engine="openpyxl")
        self.workbook = self.writer.book

    def parse_args(self) -> argparse.Namespace:
        """
        Parse command line arguments

        Returns
        -------
        args : Namespace
            Namespace of passed command line argument inputs
        """
        parser = argparse.ArgumentParser()
        parser.add_argument(
            "--output", "-o", required=True, help="output xlsm file name"
        )
        parser.add_argument("-html", required=True, help="html input")
        parser.add_argument(
            "--variant", "-v", required=True, help="variant csv file"
        )
        parser.add_argument(
            "--SV", "-sv", required=True, help="structural variant csv file"
        )
        parser.add_argument(
            "--cancer_gp",
            "-cg",
            required=True,
            help=(
                "ref cancer group - has to be one of these:"
                " COSMIC_Cancer_Genes, Haem, Medulloblastoma,"
                " MPNST, Neuro, Ovarian, Scarcoma"
            ),
        )

        return parser.parse_args()

    def generate(self) -> None:
        """
        Calls all methods in excel() to generate output xlsx
        """
        self.download_html_img()
        self.write_sheets()
        self.workbook.save(self.args.output)
        print("Done!")

    def download_html_img(self) -> None:
        """
        get the image links from html file
        """
        soup = self.get_soup()
        n = 1
        for link in soup.findAll("img"):
            img_link = link.get("src")
            self.download_image(img_link, "./", f"figure_{n}")
            if n == 2:
                self.crop_img("figure_2.jpg", 600, 600, 2400, 2400)
            n = n + 1

    def download_image(self, url, file_path, file_name) -> None:
        """
        Download the img from html links
        """
        full_path = file_path + file_name + ".jpg"
        urllib.request.urlretrieve(url, full_path)

    def crop_img(self, img_to_crop, left, top, right, bottom) -> None:
        """
        crop the image
        Parameters
        ---------
        str - img file name
        int - left margin to crop
        int - top margin to crop
        int - right margin to crop
        int - bottom margin to crop
        """
        im = Image.open(img_to_crop)
        im1 = im.crop((left, top, right, bottom))
        im1.save("cropped_" + img_to_crop)

    def read_html_tables(self, table_num) -> list:
        """
        get the tables from html file

        Parameters
        ----------
        int - table number

        Returns
        -------
        list of html table
        """
        soup = self.get_soup()
        tables = soup.findAll("table")
        info = tables[table_num]
        datasets = self.extract_data_from_html_table(info)

        return datasets

    def extract_data_from_html_table(self, table_info) -> list:
        """
        strip html table and return as list
        Parameters
        ----------
        bs4.element.Tag for html table table

        Returns
        -------
        list of html table
        """
        headings = [
            th.get_text() for th in table_info.find("tr").find_all("th")
        ]
        datasets = []
        for row in table_info.find_all("tr")[1:]:
            dataset = dict(
                zip(headings, (td.get_text() for td in row.find_all("td")))
            )
            datasets.append(dataset)
        return datasets

    def get_soup(self) -> BeautifulSoup:
        """
        get Beautiful soup obj from html

        Returns
        -------
        Beautiful soup object
        """
        url = self.args.html
        f = urllib.request.urlopen(url)
        page = f.read()
        f.close()
        soup = BeautifulSoup(page, features="lxml")

        return soup

    def get_pid(self) -> None:
        """
        get pid from html
        ### to remove as this function no longer needed
        """
        soup = self.get_soup()
        pid_html = soup.find("div", id="pid")
        pid_soup = BeautifulSoup(f"{pid_html}", features="lxml")
        pid = pid_soup.get_text()
        self.pt_name = re.search("Name: (.*)Date", pid).group(1)
        self.dob = re.search("Birth: (.*)NHS", pid).group(1)
        self.NHS_no = re.search("No.: (.*)", pid).group(1)

    def get_tmb(self) -> BeautifulSoup:
        """
        get tumor mutation burden (tmb) from html

        Returns
        -------
        bs4.element.NavigableString for tmb
        """
        soup = self.get_soup()
        pattern = re.compile(
            ("Total number of somatic non-synonymous"
             " small variants per megabase")
        )
        tmb = soup.find("b", text=pattern).next_sibling
        return tmb

    def write_sheets(self) -> None:
        """
        Write sheets to xlsx file
        """
        print("Writing sheets")
        self.write_refgene()
        self.get_pid()
        self.soc = self.workbook.create_sheet("SOC")
        self.write_soc()
        self.QC = self.workbook.create_sheet("QC")
        self.write_QC()
        self.plot = self.workbook.create_sheet("Plot")
        self.write_plot()
        self.signatures = self.workbook.create_sheet("Signatures")
        self.write_signatures()
        self.germline = self.workbook.create_sheet("Germline")
        self.write_germline()
        self.summary = self.workbook.create_sheet("Summary")
        self.write_summary()
        self.write_SNV()
        self.write_SV()

    def set_col_width(self, cell_width, sheet) -> None:
        """
        set the column width for given col in given sheet
        """
        for cell, width in cell_width:
            sheet.column_dimensions[cell].width = width

    def bold_cell(self, cells_to_bold, sheet) -> None:
        """
        bold the given cells in given sheet
        """
        for cell in cells_to_bold:
            sheet[cell].font = Font(bold=True, name=DEFAULT_FONT.name)

    def colour_cell(self, cells_to_colour, sheet, fill) -> None:
        """
        colour the given cells in given sheet
        """
        for cell in cells_to_colour:
            sheet[cell].fill = fill

    def all_border(self, row_ranges, sheet) -> None:
        """
        create all borders for given cells in given sheet
        """
        for row in row_ranges:
            for cells in sheet[row]:
                for cell in cells:
                    cell.border = THIN_BORDER

    def lower_border(self, cells_lower_border, sheet) -> None:
        """
        create lower cell border for given cells in given sheet
        """
        for cell in cells_lower_border:
            sheet[cell].border = LOWER_BORDER

    def write_soc(self) -> None:
        """
        Write soc sheet
        """
        self.patient_info = self.read_html_tables(0)
        # write titles
        self.soc.cell(1, 1).value = "Patient Details (Epic demographics)"
        self.soc.cell(1, 3).value = "Previous testing"
        self.soc.cell(2, 1).value = "NAME"
        self.soc.cell(2, 3).value = "Alteration"
        self.soc.cell(2, 4).value = "Assay"
        self.soc.cell(2, 5).value = "Result"
        self.soc.cell(2, 6).value = "WGS concordance"
        self.soc.cell(3, 1).value = "DOB, SEX"
        self.soc.cell(4, 1).value = "patient ID"
        self.soc.cell(5, 1).value = "MRN"
        self.soc.cell(6, 1).value = "NHS no."
        self.soc.cell(8, 1).value = "Histology"
        self.soc.cell(12, 1).value = "Comments"
        self.soc.cell(16, 1).value = "WGS in-house gene panel applied"
        self.soc.cell(17, 1).value = self.args.cancer_gp

        # merge columns that have longer text
        self.soc.merge_cells(
            start_row=1, end_row=1, start_column=3, end_column=6
        )
        # align cells
        cell_to_align = ["C1", "C2", "D2", "E2", "F2"]
        for cell in cell_to_align:
            self.soc[cell].alignment = Alignment(
                wrapText=True, horizontal="center"
            )
            if cell != "C1":
                self.soc[cell].font = Font(italic=True)

        # titles to set to bold
        to_bold = ["A1", "A8", "A12", "A16", "C1"]
        self.bold_cell(to_bold, self.soc)

        # set column widths for readability
        cell_col_width = (
            ("A", 32),
            ("C", 16),
            ("E", 16),
            ("D", 26),
            ("F", 26),
        )
        self.set_col_width(cell_col_width, self.soc)

        # colour cells
        greenFill = PatternFill(patternType="solid", start_color="90EE90")
        colour_cells = ["C3", "D3", "E3", "F3", "C4", "D4", "E4", "F4"]
        self.colour_cell(colour_cells, self.soc, greenFill)

        # set borders around table areas
        row_ranges = [
            "C1:F1",
            "C2:F2",
            "C3:F3",
            "C4:F4",
            "C5:F5",
            "C6:F6",
            "C7:F7",
            "C8:F8",
        ]
        self.all_border(row_ranges, self.soc)
        cells_lower_border = ["A1", "A8", "A12", "A16"]
        self.lower_border(cells_lower_border, self.soc)

        # add dropdowns
        cells_for_concordance = []
        for i in range(3, 17):
            cells_for_concordance.append(f"F{i}")
        concordance_options = '"Novel,Concordant (detected),Concordant \
                              (undetected),Disconcordant (detected) \
                              ,Disconcordant (undetected),N/A"'
        self.get_drop_down(
            dropdown_options=concordance_options,
            prompt="Select from the list",
            title="WGS concordance",
            sheet=self.soc,
            cells=cells_for_concordance,
        )

        cells_for_result = []
        for i in range(3, 17):
            cells_for_result.append(f"E{i}")
        result_options = '"Detected, Not detected"'
        self.get_drop_down(
            dropdown_options=result_options,
            prompt="Select from the list",
            title="Result",
            sheet=self.soc,
            cells=cells_for_result,
        )

        cells_for_assay = []
        for i in range(3, 17):
            cells_for_assay.append(f"D{i}")
        assay_options = '"FISH,IHC,NGS,Sanger,NGS multi-gene panel, \
                        RNA fusion panel,SNP array, Methylation array, \
                        MALDI-TOF, MLPA, MS-MLPA, Chromosome breakage, \
                        Digital droplet PCR, RT-PCR, LR-PCR"'
        self.get_drop_down(
            dropdown_options=assay_options,
            prompt="Select from the list",
            title="Assay",
            sheet=self.soc,
            cells=cells_for_assay,
        )

    def write_QC(self) -> None:
        """
        write QC sheet
        """
        # get QC info from html tables
        tumor_info = self.read_html_tables(1)
        sample_info = self.read_html_tables(2)
        germline_info = self.read_html_tables(3)
        seq_info = self.read_html_tables(4)
        tmb_value = self.get_tmb()
        # PID table
        self.QC.cell(1, 1).value = "=SOC!A2"
        self.QC.cell(2, 1).value = "=SOC!A3"
        self.QC.cell(3, 1).value = "=SOC!A5"
        self.QC.cell(4, 1).value = "=SOC!A6"
        self.QC.cell(6, 1).value = "=SOC!A9"
        self.QC.cell(8, 1).value = "QC alerts"
        self.QC.cell(9, 1).value = "None"

        # table 1
        table1_keys = (
            (3, "Diagnosis Date"),
            (4, "Tumour Received"),
            (5, "Tumour ID"),
            (6, "Presentation"),
            (7, "Diagnosis"),
            (8, "Tumour Site"),
            (9, "Tumour Type"),
            (10, "Germline Sample"),
        )
        for cell, key in table1_keys:
            self.QC.cell(1, cell).value = key

        table1_values = (
            (3, tumor_info[0]["Tumour Diagnosis Date"]),
            (4, sample_info[0]["Clinical Sample Date Time"]),
            (5, tumor_info[0]["Histopathology or SIHMDS LAB ID"]),
            (
                6,
                tumor_info[0]["Presentation"].split("_")[0]
                + " ("
                + tumor_info[0]["Primary or Metastatic"]
                + ")"
            ),
            (7, self.patient_info[0]["Clinical Indication"]),
            (8, tumor_info[0]["Tumour Topography"]),
            (
                9,
                sample_info[0]["Storage Medium"]
                + " "
                + sample_info[0]["Source"],
            ),
            (
                10,
                germline_info[0]["Storage Medium"]
                + " ("
                + germline_info[0]["Source"]
                + ")",
            ),
        )
        for cell, value in table1_values:
            self.QC.cell(2, cell).value = value

        # table 2
        table2_keys = (
            (3, "Purity (Histo)"),
            (4, "Purity (Calculated)"),
            (5, "Ploidy"),
            (6, "Total SNVs"),
            (7, "Total Indels"),
            (8, "Total SVs"),
            (9, "TMB"),
        )
        for cell, key in table2_keys:
            self.QC.cell(4, cell).value = key

        table2_values = (
            (3, sample_info[0]["Tumour Content"]),
            (4, sample_info[0]["Calculated Tumour Content"]),
            (5, sample_info[0]["Calculated Overall Ploidy"]),
            (6, seq_info[1]["Total somatic SNVs"]),
            (7, seq_info[1]["Total somatic indels"]),
            (8, seq_info[1]["Total somatic SVs"]),
            (9, str(tmb_value).strip()),
        )
        for cell, value in table2_values:
            self.QC.cell(5, cell).value = value

        # table 3
        table3_keys = (
            (3, "Sample type"),
            (4, "Mean depth, x"),
            (5, "Mapped reads, %"),
            (6, "Chimeric DNA frag, %"),
            (7, "Insert size, bp"),
            (8, "Unevenness, x"),
        )
        for cell, key in table3_keys:
            self.QC.cell(7, cell).value = key

        seq_info_title = (
            (3, "Sample type"),
            (4, "Genome-wide coverage mean, x"),
            (5, "Mapped reads, %"),
            (6, "Chimeric DNA fragments, %"),
            (7, "Insert size median, bp"),
            (8, "Unevenness of local genome coverage, x"),
        )
        for cell, title in seq_info_title:
            self.QC.cell(8, cell).value = seq_info[0][title]
            self.QC.cell(9, cell).value = seq_info[1][title]

        # titles to set to bold
        to_bold = [
            "A1",
            "A8",
            "C1",
            "D1",
            "E1",
            "F1",
            "G1",
            "H1",
            "I1",
            "J1",
            "C4",
            "D4",
            "E4",
            "F4",
            "G4",
            "H4",
            "I4",
            "C7",
            "D7",
            "E7",
            "F7",
            "G7",
            "H7",
        ]
        self.bold_cell(to_bold, self.QC)

        # set column widths for readability
        cell_col_width = (
            ("A", 32),
            ("B", 8),
            ("C", 22),
            ("D", 22),
            ("E", 22),
            ("F", 22),
            ("G", 22),
            ("H", 22),
            ("I", 22),
            ("J", 22),
        )
        self.set_col_width(cell_col_width, self.QC)

        # set borders around table areas
        row_ranges = [
            "C1:J1",
            "C2:J2",
            "C4:I4",
            "C5:I5",
            "C7:H7",
            "C8:H8",
            "C9:H9",
        ]
        self.all_border(row_ranges, self.QC)
        self.lower_border(["A8"], self.QC)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="ADD8E6")
        blue_colour_cells = ["C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1",
                             "C4", "D4", "E4", "F4", "G4", "H4", "I4",
                             "C7", "D7", "E7", "F7", "G7", "H7"]
        self.colour_cell(blue_colour_cells, self.QC, blueFill)

        # add dropdowns
        cells_for_QC = ["A9"]
        QC_options = '"None,<30% tumour purity,SNVs low VAF (<6%),TINC (<5%)"'
        self.get_drop_down(
            dropdown_options=QC_options,
            prompt="Select from the list",
            title="QC alerts",
            sheet=self.QC,
            cells=cells_for_QC,
        )
        # insert img from html
        self.insert_img(self.QC, "figure_9.jpg", "C12", 400, 600)
        self.insert_img(self.QC, "figure_11.jpg", "G12", 400, 600)

    def write_plot(self) -> None:
        """
        write plot sheet
        """
        # pid table
        self.plot.cell(1, 1).value = "=SOC!A2"
        self.plot.cell(2, 1).value = "=SOC!A3"
        self.plot.cell(3, 1).value = "=SOC!A5"
        self.plot.cell(4, 1).value = "=SOC!A6"
        self.plot.cell(6, 1).value = "=SOC!A9"
        self.plot.cell(8, 1).value = "Pertinent chromosomal CNVs"
        self.plot.cell(9, 1).value = "None"
        self.plot.cell(5, 4).value = "Insert"

        # titles to set to bold
        to_bold = ["A1", "A8"]
        self.bold_cell(to_bold, self.plot)

        # set column widths for readability
        self.plot.column_dimensions["A"].width = 32

        # insert img from html file
        self.insert_img(self.plot, "cropped_figure_2.jpg", "C2", 500, 500)
        self.insert_img(self.plot, "figure_3.jpg", "C26", 600, 1200)

    def write_signatures(self) -> None:
        """
        write signatures sheet
        """
        # pid table
        self.signatures.cell(1, 1).value = "=SOC!A2"
        self.signatures.cell(2, 1).value = "=SOC!A3"
        self.signatures.cell(3, 1).value = "=SOC!A5"
        self.signatures.cell(4, 1).value = "=SOC!A6"
        self.signatures.cell(6, 1).value = "=SOC!A9"
        self.signatures.cell(8, 1).value = "Signature version"
        self.signatures.cell(9, 1).value = "v2 (March 2015)"
        self.signatures.cell(13, 1).value = "Pertinent signatures"
        self.signatures.cell(14, 1).value = "None"

        # titles to set to bold
        to_bold = ["A1", "A8", "A13"]
        self.bold_cell(to_bold, self.signatures)
        # set lower border
        self.lower_border(["A8"], self.signatures)

        # set column widths for readability
        self.signatures.column_dimensions["A"].width = 32

        # insert img from html file
        self.insert_img(self.signatures, "figure_6.jpg", "C3", 600, 800)
        self.insert_img(self.signatures, "figure_7.jpg", "N3", 600, 800)

    def get_clnsigconf(self, clinvarID) -> str:
        """
        get the clnsigconf from clinvar file for given
        clinvar ID

        Parameters
        ----------
        int for clinvar ID
        str for clinvar ref file

        Returns
        -------
        list for CLNSIG, CLNSIGCONF
        """
        clinvar_dx = []
        for c in ["CLNSIG", "CLNSIGCONF"]:
            cmd = f"zcat {CLINVAR_REF} | awk '$3=={clinvarID} \
                  {{print($8)}}' |  \
                  grep -o -P '(?<={c}=).*?(?=;)'"
            ps = subprocess.Popen(
                cmd, shell=True, stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT
            )
            output = ps.communicate()[0]
            clinvar_dx.append(output.decode("utf-8").strip())
        if all('' == s for s in clinvar_dx):
            clinvar_dx = ""        
        elif '' in clinvar_dx:
            clinvar_dx = next(s for s in clinvar_dx if s)
        else:
            clinvar_dx = clinvar_dx[1]
        return clinvar_dx

    def write_germline(self) -> None:
        """
        write germline sheet
        """
        self.germline.cell(1, 1).value = "=SOC!A2"
        self.germline.cell(2, 1).value = "=SOC!A3"
        self.germline.cell(3, 1).value = "=SOC!A5"
        self.germline.cell(4, 1).value = "=SOC!A6"
        self.germline.cell(6, 1).value = "=SOC!A9"
        self.germline.cell(8, 1).value = "Pertinent germline variants"
        self.germline.cell(9, 1).value = "None"

        # Germline SNV table
        self.germline.cell(1, 3).value = "Germline SNV"
        snv_table_keys = (
            (3, "Gene"),
            (4, "GRCh38 Coordinates"),
            (5, "Variant"),
            (6, "Genotype"),
            (7, "Role in Cancer"),
            (8, "ClinVar"),
            (9, "gnomAD"),
            (10, "Tumour VAF"),
        )
        for cell, key in snv_table_keys:
            self.germline.cell(2, cell).value = key

        # populate germline table
        germline_table = pd.read_csv(self.args.variant, sep=",")
        germline_table = germline_table[germline_table["Origin"] == "germline"]
        germline_table.reset_index(drop=True, inplace=True)

        # get the clnsigconf from clinvar file based on clinvar ID
        clinvarID = list(germline_table["ClinVar ID"])
        d = []
        for cid in clinvarID:
            d.append(
                {
                    "ClinVar ID": cid,
                    "clnsigconf": self.get_clnsigconf(cid),
                }
            )
        clinvar_df = pd.DataFrame(d)

        germline_table = germline_table.merge(
            clinvar_df, on="ClinVar ID", how="left"
        )
        germline_table = germline_table[
            [
                "Gene",
                "GRCh38 coordinates;ref/alt allele",
                "CDS change and protein change",
                "Genotype",
                "Gene mode of action",
                "clnsigconf",
                "Population germline allele frequency (GE | gnomAD)",
            ]
        ]
        # split the col to get gnomAD
        germline_table[["GE", "gnomAD"]] = germline_table[
            "Population germline allele frequency (GE | gnomAD)"
        ].str.split("|", expand=True)
        germline_table.drop(
            ["GE", "Population germline allele frequency (GE | gnomAD)"],
            axis=1,
            inplace=True,
        )

        # write df into excel sheet
        num_gene = germline_table.shape[0]
        rows = dataframe_to_rows(germline_table)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if c_idx != 1 and r_idx != 1:
                    self.germline.cell(
                        row=r_idx, column=c_idx + 1, value=value
                    )

        self.germline.cell(
            num_gene + 4, 3
        ).value = "Clinical genetics feedback"

        # titles to set to bold
        to_bold = [
            "A1",
            "A8",
            "C1",
            "C2",
            "D2",
            "E2",
            "F2",
            "G2",
            "H2",
            "I2",
            "J2",
            f"C{num_gene+4}",
        ]
        self.bold_cell(to_bold, self.germline)
        # set border
        cells_lower_border = [
            "A8",
            f"C{num_gene+4}",
        ]
        self.lower_border(cells_lower_border, self.germline)

        # set column widths for readability
        self.germline.column_dimensions["A"].width = 32

        # set borders around table areas
        row_ranges = []
        for i in range(1, num_gene + 3):
            row_ranges.append(f"C{i}:J{i}")
        self.all_border(row_ranges, self.germline)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="ADD8E6")
        blue_colour_cells = ["C2", "D2", "E2", "F2", "G2", "H2", "I2", "J2"]

        self.colour_cell(blue_colour_cells, self.germline, blueFill)

        # set column widths for readability
        cell_col_width = (
            ("A", 36),
            ("B", 10),
            ("C", 28),
            ("D", 32),
            ("E", 32),
            ("F", 20),
            ("G", 28),
            ("H", 40),
            ("I", 20),
            ("J", 28),
        )
        self.set_col_width(cell_col_width, self.germline)

    def write_summary(self) -> None:
        """
        Write summary sheet
        """
        # pid table
        self.summary.cell(1, 1).value = "=SOC!A2"
        self.summary.cell(2, 1).value = "=SOC!A3"
        self.summary.cell(3, 1).value = "=SOC!A5"
        self.summary.cell(4, 1).value = "=SOC!A6"
        self.summary.cell(6, 1).value = "=SOC!A9"
        self.summary.cell(9, 1).value = "Reportable genes"
        self.summary.cell(
            10, 1
        ).value = '= _xlfn.TEXTJOIN(", ",TRUE,C21:C28,C33:C40)'
        self.summary.cell(12, 1).value = "Comments"

        # snv table
        self.summary.cell(19, 3).value = "Somatic SNV"
        snv_table_keys = (
            (3, "Gene"),
            (4, "GRCh38 Coordinates"),
            (5, "Mutation"),
            (6, "VAF"),
            (7, "Variant Class"),
            (8, "Actionability"),
        )
        for cell, key in snv_table_keys:
            self.summary.cell(20, cell).value = key
        # add formula
        for row in range(21, 29):
            ref_row = row + 22
            for col in ["C", "D", "E", "F"]:
                self.summary[f"{col}{row}"] = f"=summary!{col}{ref_row}"

        # cnv sv table
        self.summary.cell(30, 3).value = "Somatic CNV_SV"
        cnv_sv_table_key = (
            (3, "Gene/Locus"),
            (4, "GRCh38 Coordinates"),
            (5, "Cytological Bands"),
            (6, "Variant Type"),
            (7, "Variant Class"),
            (8, "Actionability"),
        )
        for cell, key in cnv_sv_table_key:
            self.summary.cell(31, cell).value = key
        # add formula
        for row in range(32, 40):
            ref_row = row + 21
            for col in ["C", "D", "E", "F"]:
                self.summary[f"{col}{row}"] = f"=summary!{col}{ref_row}"

        self.summary.cell(41, 1).value = "SNV"
        # snv title
        snv_title_keys = (
            (1, "Domain"),
            (2, "Origin"),
            (3, "Gene"),
            (4, "GRCh38 coordinates;ref/alt allele"),
            (5, "Transcript"),
            (6, "CDS change and protein change"),
            (7, "Predicted consequences"),
            (8, "Population germline allele frequency (GE | gnomAD)"),
            (9, "VAF"),
            (10, "Alt allele/total read depth"),
            (11, "Genotype"),
            (12, "COSMIC ID"),
            (13, "ClinVar ID"),
            (14, "ClinVar review status"),
            (15, "ClinVar clinical significance"),
            (16, "Gene mode of action"),
            (17, "Recruiting Clinical Trials 30 Jan 2023"),
            (18, "PharmGKB_ID"),
        )
        for cell, key in snv_title_keys:
            self.summary.cell(42, cell).value = key

        self.summary.cell(51, 1).value = "CNV_SV"

        # cnv_sv title
        cnv_sv_title_keys = (
            (1, "Origin"),
            (2, "Variant domain"),
            (3, "Event domain"),
            (4, "Gene"),
            (5, "Transcript"),
            (6, "Impacted transcript region"),
            (7, "GRCh38 coordinates"),
            (8, "Type"),
            (9, "Size"),
            (
                10,
                ("Population germline allele frequency"
                 " (GESG | GECG for somatic SVs or AF |"
                 " AUC for germline CNVs)"),
            ),
            (11, "Confidence/support"),
            (12, "Chromosomal bands"),
            (13, "Recruiting Clinical Trials 30 Jan 2023"),
            (14, "ClinVar clinical significance"),
            (15, "Gene mode of action"),
        )
        for cell, key in cnv_sv_title_keys:
            self.summary.cell(52, cell).value = key

        # titles to set to bold
        to_bold = [
            "A1",
            "A9",
            "A12",
            "A16",
            "C19",
            "C20",
            "C30",
            "D20",
            "E20",
            "F20",
            "G20",
            "H20",
            "C31",
            "D31",
            "E31",
            "F31",
            "G31",
            "H31",
            "A41",
            "A51",
        ]
        self.bold_cell(to_bold, self.summary)

        # set column widths for readability
        cell_col_width = (
            ("A", 32),
            ("C", 22),
            ("D", 26),
            ("F", 26),
            ("G", 26),
            ("H", 26),
        )
        self.set_col_width(cell_col_width, self.summary)

        # colour title cells
        blueFill = PatternFill(patternType="solid", start_color="90EE90")

        colour_cells = [
            "C20",
            "D20",
            "E20",
            "F20",
            "G20",
            "H20",
            "C31",
            "D31",
            "E31",
            "F31",
            "G31",
            "H31",
        ]
        self.colour_cell(colour_cells, self.summary, blueFill)

        # set borders around table areas
        row_ranges = [
            "C20:H20",
            "C21:H21",
            "C22:H22",
            "C23:H23",
            "C24:H24",
            "C25:H25",
            "C26:H26",
            "C27:H27",
            "C28:H28",
            "C31:H31",
            "C32:H32",
            "C33:H33",
            "C34:H34",
            "C35:H35",
            "C36:H36",
            "C37:H37",
            "C38:H38",
            "C39:H39",
        ]
        self.all_border(row_ranges, self.summary)
        cells_lower_border = ["A9", "A12", "A41", "A51"]
        self.lower_border(cells_lower_border, self.summary)

        smaller_font = Font(size=8)
        for i in range(41, 72):
            for cell in self.summary[f"{i}:{i}"]:
                cell.font = smaller_font

        # add dropdowns
        cells_for_class = []
        for i in range(21, 40):
            if i not in [29, 30, 31]:
                cells_for_class.append(f"G{i}")

        class_options = (
            '"Pathogenic, Likely pathogenic, Uncertain, Likely benign, Benign"'
        )
        self.get_drop_down(
            dropdown_options=class_options,
            prompt="Select from the list",
            title="Variant class",
            sheet=self.summary,
            cells=cells_for_class,
        )

        cells_for_action = []
        for i in range(21, 40):
            if i not in [29, 30, 31]:
                cells_for_action.append(f"H{i}")

        action_options = ('"1. Predicts therapeutic response,'
                          ' 2. Prognostic, 3. Defines diagnosis group'
                          ', 4. Eligibility for trial, 5. Other"')
        self.get_drop_down(
            dropdown_options=action_options,
            prompt="Select from the list",
            title="Actionability",
            sheet=self.summary,
            cells=cells_for_action,
        )

        # insert img from html file
        self.insert_img(self.summary, "cropped_figure_2.jpg", "C1", 350, 350)
        self.insert_img(self.summary, "figure_3.jpg", "E1", 300, 700)

    def write_refgene(self) -> None:
        """
        write RefGene sheet
        """
        self.df_rgg = pd.read_csv(REFGENEGP_REF)
        self.df_refgene = self.df_rgg[
            self.df_rgg["RefGene Group"] == self.args.cancer_gp
        ]
        self.df_refgene.drop_duplicates(
            subset="Gene", keep="last", inplace=True
        )  # TO DO: TO REMOVE##should be corrected for ovarian and medullo
        self.df_refgene.reset_index(drop=True, inplace=True)
        self.df_refgene.to_excel(
            self.writer, sheet_name="RefGene", index=False
        )
        ref_gene = self.writer.sheets["RefGene"]
        filters = ref_gene.auto_filter
        filters.ref = "A:G"

    def write_SNV(self) -> None:
        """
        write SNV sheet
        """
        self.df_hotspots = pd.read_csv(HTOSPOTS_REF)
        df = pd.read_csv(self.args.variant, sep=",")
        # select only somatic rows
        df = df[df["Origin"] == "somatic"]
        df.reset_index(drop=True, inplace=True)
        num_variant = df.shape[0]
        df["Report (Y/N)"] = ""
        df["Comments"] = ""
        df[["c_dot", "p_dot"]] = df["CDS change and protein change"].str.split(
            r"(?=p)", n=1, expand=True
        )
        df["c_dot"] = df["c_dot"].str.replace("(;$)", "", regex=True)
        df["Alteration_RefGene"] = df["Gene"].map(
            self.df_refgene.set_index("Gene")["Alteration"]
        )
        # look up genes from df_refgene
        df["Origin_RefGene"] = df["Gene"].map(
            self.df_refgene.set_index("Gene")["Origin"]
        )
        df["Entities_RefGene"] = df["Gene"].map(
            self.df_refgene.set_index("Gene")["Entities"]
        )
        df["Comments_RefGene"] = df["Gene"].map(
            self.df_refgene.set_index("Gene")["Comments"]
        )
        df = df.replace([None], [""], regex=True)
        df["MTBP c."] = df["Gene"] + ":" + df["c_dot"]
        df["MTBP p."] = df["Gene"] + ":" + df["p_dot"]
        df[["HS p.", "col1", "col2"]] = df["MTBP p."].str.split(
            r"([^\d]+)$", expand=True
        )
        df.drop(["col1", "col2"], axis=1, inplace=True)

        df["HS_Sample"] = df["HS p."].map(
            self.df_hotspots.set_index("HS_PROTEIN_ID")["HS_Samples"]
        )
        df["HS_Tumour"] = df["HS p."].map(
            self.df_hotspots.set_index("HS_PROTEIN_ID")[
                "HS_Tumor Type Composition"
            ]
        )
        df[["GE", "gnomAD"]] = df[
            "Population germline allele frequency (GE | gnomAD)"
        ].str.split("|", expand=True)
        df.loc[:, "Variant_to_report"] = ""
        df = df[
            [
                "Gene",
                "Origin",
                "GRCh38 coordinates;ref/alt allele",
                "CDS change and protein change",
                "Predicted consequences",
                "gnomAD",
                "VAF",
                "Genotype",
                "COSMIC ID",
                "Gene mode of action",
                "Report (Y/N)",
                "Comments",
                "Alteration_RefGene",
                "Origin_RefGene",
                "Entities_RefGene",
                "Comments_RefGene",
                "MTBP c.",
                "MTBP p.",
                "HS_Sample",
                "HS_Tumour",
                "Variant_to_report",
            ]
        ]
        df.rename(
            columns={
                "GRCh38 coordinates;ref/alt allele": "GRCh38 coordinates",
                "CDS change and protein change": "Variant",
            },
            inplace=True,
        )
        df.to_excel(self.writer, sheet_name="SNV", index=False)
        self.SNV = self.writer.sheets["SNV"]
        cell_col_width = (
            ("A", 12),
            ("B", 12),
            ("C", 28),
            ("D", 28),
            ("E", 14),
            ("F", 14),
            ("G", 8),
            ("H", 12),
            ("I", 14),
            ("J", 20),
            ("K", 20),
            ("L", 26),
            ("M", 26),
            ("N", 20),
            ("O", 14),
            ("P", 22),
            ("Q", 26),
            ("R", 18),
            ("S", 18),
            ("T", 16),
            ("U", 16)
        )
        self.set_col_width(cell_col_width, self.SNV)

        # get max col for dropdown
        max_col = df.shape[1]
        max_col_letter = get_column_letter(max_col)
        # add filter in col
        filters = self.SNV.auto_filter
        filters.ref = f"A:{max_col_letter}"

        report_options = '"yes, no"'
        cells_for_report = []
        for i in range(2, num_variant + 2):
            cells_for_report.append(f"{max_col_letter}{i}")

        self.get_drop_down(
            dropdown_options=report_options,
            prompt="Select from the list",
            title="yes or no",
            sheet=self.SNV,
            cells=cells_for_report,
        )

    def write_SV(self) -> None:
        """
        write SV, SV_loss and SV_gain sheets
        """
        df_SV = pd.read_csv(self.args.SV, sep=",")
        df_SV["gene_count"] = df_SV["Gene"].str.count(r"\;")
        max_num_gene = df_SV["gene_count"].max() + 1
        # split gene col and create look up col for them
        if max_num_gene == 1:
            df_SV["A_Gene"] = df_SV["Gene"]

        elif max_num_gene == 2:
            df_SV[["A_Gene", "B_Gene"]] = df_SV["Gene"].str.split(
                ";", expand=True
            )
            df_SV["B_LOOKUP"] = np.where(
                df_SV["B_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["B_Gene"],
                "",
            )

        elif max_num_gene == 3:
            df_SV[["A_Gene", "B_Gene", "C_Gene"]] = df_SV["Gene"].str.split(
                ";", expand=True
            )
            df_SV["B_LOOKUP"] = np.where(
                df_SV["B_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["B_Gene"],
                "",
            )
            df_SV["C_LOOKUP"] = np.where(
                df_SV["C_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["C_Gene"],
                "",
            )
        elif max_num_gene == 4:
            df_SV[["A_Gene", "B_Gene", "C_gene", "D_gene"]] = df_SV[
                "Gene"
            ].str.split(";", expand=True)
            df_SV["B_LOOKUP"] = np.where(
                df_SV["B_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["B_Gene"],
                "",
            )
            df_SV["C_LOOKUP"] = np.where(
                df_SV["C_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["C_Gene"],
                "",
            )
            df_SV["D_LOOKUP"] = np.where(
                df_SV["D_Gene"].isin(list(self.df_refgene["Gene"])),
                df_SV["D_Gene"],
                "",
            )
        df_SV["Report (Y/N)"] = ""
        df_SV["Comments"] = ""
        # look up A_Gene in df_refgene
        df_SV["Alteration_RefGene"] = df_SV["A_Gene"].map(
            self.df_refgene.set_index("Gene")["Alteration"]
        )
        df_SV["Origin_RefGene"] = df_SV["A_Gene"].map(
            self.df_refgene.set_index("Gene")["Origin"]
        )
        df_SV["Entities_RefGene"] = df_SV["A_Gene"].map(
            self.df_refgene.set_index("Gene")["Entities"]
        )
        df_SV["Comments_RefGene"] = df_SV["A_Gene"].map(
            self.df_refgene.set_index("Gene")["Comments"]
        )
        df_SV["Comments_RefGene"] = df_SV["A_Gene"].map(
            self.df_refgene.set_index("Gene")["Comments"]
        )

        # subset df
        lookup_col = [col for col in df_SV if col.endswith("LOOKUP")]
        selected_col = [
            "Variant domain",
            "Gene",
            "GRCh38 coordinates",
            "Type",
            "Gene mode of action",
            "Report (Y/N)",
            "Comments",
            "Alteration_RefGene",
            "Origin_RefGene",
            "Entities_RefGene",
            "Comments_RefGene",
        ] #+ lookup_col
        df_SV = df_SV[selected_col]
        # split df into three df
        df_loss = df_SV[df_SV["Type"].str.lower().str.contains("loss|loh")]
        df_loss[["Type", "Type_num"]] = df_loss.Type.str.split(
            r"\(|\)", expand=True
        ).iloc[:, [0, 1]]
        df_gain = df_SV[df_SV["Type"].str.lower().str.contains("gain")]
        df_gain[["Type", "Type_num"]] = df_gain.Type.str.split(
            r"\(|\)", expand=True
        ).iloc[:, [0, 1]]
        df_other = df_SV[
            ~df_SV["Type"].str.lower().str.contains("loss|loh|gain")
        ]
        df_other[["Type1", "Type2"]] = df_other.Type.str.split(
            ";", expand=True
        )
        df_to_write = (
            (df_loss, "SV_loss"),
            (df_gain, "SV_gain"),
            (df_other, "SV_others"),
        )
        # write each df into sheet
        for df, sheet_name in df_to_write:
            num_variant = df.shape[0]

            if sheet_name == "SV_others":
                
                reordered_col = [
                    "Variant domain",
                    "Gene",
                    "GRCh38 coordinates",
                    "Type1",
                    "Type2",
                    "Gene mode of action",
                    "Report (Y/N)",
                    "Comments",
                    "Alteration_RefGene",
                    "Origin_RefGene",
                    "Entities_RefGene",
                    "Comments_RefGene",
                ] #+ lookup_col

                df = df[reordered_col]
                df.loc[:, "Variant_to_report"] = ""
            else:
                reordered_col = [
                    "Variant domain",
                    "Gene",
                    "GRCh38 coordinates",
                    "Type",
                    "Type_num",
                    "Gene mode of action",
                    "Report (Y/N)",
                    "Comments",
                    "Alteration_RefGene",
                    "Origin_RefGene",
                    "Entities_RefGene",
                    "Comments_RefGene",
                ] #+ lookup_col
                df = df[reordered_col]
                df.loc[:, "Variant_to_report"] = ""
            max_col = df.shape[1]
            df.to_excel(self.writer, sheet_name=sheet_name, index=False)
            sheet = self.writer.sheets[sheet_name]
            cell_col_width = (
                ("A", 12),
                ("B", 12),
                ("C", 22),
                ("F", 20),
                ("G", 20),
                ("H", 24),
                ("I", 24),
                ("J", 24),
                ("K", 20),
                ("L", 20)
            )

            self.set_col_width(cell_col_width, sheet)
            max_col_letter = get_column_letter(max_col)
            filters = sheet.auto_filter
            filters.ref = f"A:{max_col_letter}"
            report_options = '"yes, no"'
            cells_for_report = []
            for i in range(2, num_variant + 2):
                cells_for_report.append(f"{max_col_letter}{i}")

            self.get_drop_down(
                dropdown_options=report_options,
                prompt="Select from the list",
                title="yes or no",
                sheet=sheet,
                cells=cells_for_report,
            )

    def get_drop_down(
        self, dropdown_options, prompt, title, sheet, cells
    ) -> None:
        """
        create the drop-downs items for designated cells

        Parameters
        ----------
        dropdown_options: str
            str containing drop-down items
        prompt: str
            prompt message for drop-down
        title: str
            title message for drop-down
        sheet: openpyxl.Writer writer object
            current worksheet
        cells: list
            list of cells to add drop-down
        """
        options = dropdown_options
        val = DataValidation(type="list", formula1=options, allow_blank=True)
        val.prompt = prompt
        val.promptTitle = title
        sheet.add_data_validation(val)
        for cell in cells:
            val.add(sheet[cell])
        val.showInputMessage = True
        val.showErrorMessage = True

    def insert_img(self, sheet, img_to_insert, cell_to_insert, h, w) -> None:
        """
        insert the img downloaded from html into spreadsheet
        """
        ws = sheet
        img = drawing.image.Image(img_to_insert)
        img.height = h
        img.width = w
        img.anchor = cell_to_insert
        ws.add_image(img)


def main():
    # generate output Excel file
    excel_handler = excel()
    excel_handler.generate()


if __name__ == "__main__":
    main()
