import string

from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc


THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


CONFIG = {
    "cells_to_write": {
        (1, i): value
        for i, value in enumerate(
            [
                "Event domain",
                "Gene",
                "RefSeq IDs",
                "Impacted transcript region",
                "GRCh38 coordinates",
                "Type",
                "Copy Number",
                "Size",
                "Cyto 1",
                "Cyto 2",
                "Gene mode of action",
                "Variant class",
                "OG_Amp",
                "Focality",
                "Full transcript",
                "COSMIC Driver",
                "COSMIC Alterations",
                "Paed Driver",
                "Paed Entities",
                "Sarc Driver",
                "Sarc Entities",
                "Neuro Driver",
                "Neuro Entities",
                "Ovary Driver",
                "Ovary Entities",
                "Haem Driver",
                "Haem Entities",
            ],
            1,
        )
    },
    "to_bold": [f"{misc.convert_index_to_letters(i)}1" for i in range(0, 27)],
    "col_width": [
        ("B", 12),
        ("C", 16),
        ("D", 16),
        ("E", 22),
        ("H", 14),
        ("I", 10),
        ("J", 10),
    ],
    "borders": {
        "cell_rows": [
            ("A1:AA1", THIN_BORDER),
        ],
    },
    "row_height": [(1, 80)],
    "auto_filter": "A:AA",
    "freeze_panes": "F1",
    "text_orientation": [
        (f"{misc.convert_index_to_letters(i)}1", 90) for i in range(11, 38)
    ],
}


def add_dynamic_values(data: pd.DataFrame) -> dict:
    """Add dynamic values for the Gain config

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe for the Gain structural variants

    Returns
    -------
    dict
        Dict populated with the dynamic values processed for the Gain
        structural variants
    """

    nb_sv_variants = data.shape[0]

    config_with_dynamic_values = {
        "cells_to_write": {
            # remove the col and row index from the writing?
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(data), 1)
            for c_idx, value in enumerate(row, 1)
            if c_idx != 1 and r_idx != 1
        },
        "cells_to_colour": [
            (
                f"{col}{i}",
                PatternFill(patternType="solid", start_color="FFDBBB"),
            )
            for col in ["L", "M", "N", "O"]
            for i in range(1, nb_sv_variants + 2)
        ]
        + [
            (
                # letters P to AA
                f"{misc.convert_index_to_letters(i)}{j}",
                PatternFill(patternType="solid", start_color="c4d9ef"),
            )
            for i in range(15, 27)
            for j in range(1, nb_sv_variants + 2)
        ],
        "to_align": [f"G{i}" for i in range(2, nb_sv_variants + 2)],
        "dropdowns": [
            {
                "cells": {
                    (f"L{i}" for i in range(2, nb_sv_variants + 2)): (
                        '"Oncogenic, Likely oncogenic,'
                        "Uncertain, Likely passenger,"
                        'Likely artefact"'
                    ),
                },
                "title": "Variant class",
            },
        ],
    }

    return config_with_dynamic_values
