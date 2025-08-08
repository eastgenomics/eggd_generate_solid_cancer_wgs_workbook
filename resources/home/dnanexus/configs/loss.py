from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc

THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT_BORDER = Border(left=THIN)


CONFIG = {
    "cells_to_write": {
        (1, i): value
        for i, value in enumerate(
            [
                "Event domain",
                "Gene",
                "RefSeq IDs",
                "Impacted transcript region",
                "GRCh38 coordinates",
                "Type",
                "Copy Number",
                "Size",
                "Cyto 1",
                "Cyto 2",
                "Gene mode of action",
                "Variant class",
                "Comments",
                "TSG_Hom",
                "SNV_LOH",
                "COSMIC Driver",
                "COSMIC Entities",
                "Paed Driver",
                "Paed Entities",
                "Sarc Driver",
                "Sarc Entities",
                "Neuro Driver",
                "Neuro Entities",
                "Ovary Driver",
                "Ovary Entities",
                "Haem Driver",
                "Haem Entities",
            ],
            1,
        )
    },
    "to_bold": [f"{misc.convert_index_to_letters(i)}1" for i in range(0, 27)],
    "col_width": [
        ("A", 10),
        ("B", 12),
        ("C", 16),
        ("D", 16),
        ("E", 22),
        ("F", 6),
        ("G", 5),
        ("H", 14),
        ("I", 10),
        ("J", 10),
        ("K", 22),
        ("N", 6),
        ("O", 6),
    ]
    + [(f"{misc.convert_index_to_letters(i)}", 5) for i in range(14, 27)],
    "cells_to_colour": [
        (
            f"{col}1",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for col in ["L", "M", "N", "O"]
    ]
    + [
        (
            # letters P to AA
            f"{misc.convert_index_to_letters(i)}1",
            PatternFill(patternType="solid", start_color="fdeada"),
        )
        for i in range(15, 27)
    ],
    "borders": {
        "cell_rows": [
            ("A1:AA1", THIN_BORDER),
        ],
    },
    "row_height": [(1, 80)],
    "auto_filter": "A:AA",
    "freeze_panes": "H1",
    "alignment_info": [
        (
            f"{misc.convert_index_to_letters(i)}1",
            {
                "horizontal": "left",
                "vertical": "bottom",
                "wrapText": True,
                "text_rotation": 90,
            },
        )
        for i in range(0, 27)
    ],
}


def add_dynamic_values(data: pd.DataFrame) -> dict:
    """Add dynamic values for the Loss config

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe for the Loss structural variants

    Returns
    -------
    dict
        Dict populated with the dynamic values processed for the Loss
        structural variants
    """

    if data is None:
        return {}

    nb_sv_variants = data.shape[0]

    config_with_dynamic_values = {
        "cells_to_write": {
            # remove the col and row index from the writing?
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(data), 1)
            for c_idx, value in enumerate(row, 1)
            if c_idx != 1 and r_idx != 1
        },
        "alignment_info": [
            (f"G{i}", {"horizontal": "center"})
            for i in range(2, nb_sv_variants + 2)
        ],
        "dropdowns": [
            {
                "cells": {
                    (f"L{i}" for i in range(2, nb_sv_variants + 2)): (
                        '"Oncogenic, Likely oncogenic,'
                        "Uncertain, Likely passenger,"
                        'Likely artefact"'
                    ),
                },
                "title": "Variant class",
            },
        ],
        "borders": {
            "cell_rows": [
                (f"P1:P{nb_sv_variants+1}", LEFT_BORDER),
                (f"R1:R{nb_sv_variants+1}", LEFT_BORDER),
                (f"T1:T{nb_sv_variants+1}", LEFT_BORDER),
                (f"V1:V{nb_sv_variants+1}", LEFT_BORDER),
                (f"X1:X{nb_sv_variants+1}", LEFT_BORDER),
                (f"Z1:Z{nb_sv_variants+1}", LEFT_BORDER),
                (f"AA1:AA{nb_sv_variants+1}", LEFT_BORDER),
                (f"AB1:AB{nb_sv_variants+1}", LEFT_BORDER),
            ],
        },
    }

    return config_with_dynamic_values
