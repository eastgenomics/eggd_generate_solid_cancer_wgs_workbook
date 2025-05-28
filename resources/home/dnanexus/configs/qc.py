from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill

from utils import html


# prepare formatting
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)

CONFIG = {
    "cells_to_write": {
        (4, 1): "Diagnosis Date",
        (4, 2): "Tumour Received",
        (4, 3): "Tumour ID",
        (4, 4): "Presentation",
        (4, 5): "Diagnosis",
        (4, 6): "Tumour Site",
        (4, 7): "Tumour Type",
        (4, 8): "Germline Sample",
        (5, 1): ("Tumor info", 0, "Tumour Diagnosis Date"),
        (5, 2): ("Sample info", 0, "Clinical Sample Date Time"),
        (5, 3): ("Tumor info", 0, "Histopathology or SIHMDS LAB ID"),
        (5, 4): [
            ("Tumor info", 0, "Presentation", "split"),
            ("Tumor info", 0, "Primary or Metastatic", "parentheses"),
        ],
        (5, 5): ("Patient info", 0, "Clinical Indication"),
        (5, 6): ("Tumor info", 0, "Tumour Topography"),
        (5, 7): [
            ("Sample info", 0, "Storage Medium", ""),
            ("Sample info", 0, "Source", ""),
        ],
        (5, 8): [
            ("Germline info", 0, "Storage Medium", ""),
            ("Germline info", 0, "Source", "parentheses"),
        ],
        (7, 1): "Purity (Histo)",
        (7, 2): "Purity (Calculated)",
        (7, 3): "Ploidy",
        (7, 4): "Total SNVs",
        (7, 5): "Total Indels",
        (7, 6): "Total SVs",
        (7, 7): "TMB",
        (8, 1): ("Sample info", 0, "Tumour Content"),
        (8, 2): ("Sample info", 0, "Calculated Tumour Content"),
        (8, 3): ("Sample info", 0, "Calculated Overall Ploidy"),
        (8, 4): ("Sequencing info", 1, "Total somatic SNVs"),
        (8, 5): ("Sequencing info", 1, "Total somatic indels"),
        (8, 6): ("Sequencing info", 1, "Total somatic SVs"),
        (8, 7): html.get_tag_sibling,
        (10, 1): "Sample type",
        (10, 2): "Mean depth, x",
        (10, 3): "Mapped reads, %",
        (10, 4): "Chimeric DNA frag, %",
        (10, 5): "Insert size, bp",
        (10, 6): "Unevenness, x",
        (11, 1): ("Sequencing info", 0, "Sample type"),
        (11, 2): (
            "Sequencing info",
            0,
            "Genome-wide coverage mean, x",
        ),
        (11, 3): ("Sequencing info", 0, "Mapped reads, %"),
        (11, 4): ("Sequencing info", 0, "Chimeric DNA fragments, %"),
        (11, 5): ("Sequencing info", 0, "Insert size median, bp"),
        (11, 6): (
            "Sequencing info",
            0,
            "Unevenness of local genome coverage, x",
        ),
        (12, 1): ("Sequencing info", 1, "Sample type"),
        (12, 2): (
            "Sequencing info",
            1,
            "Genome-wide coverage mean, x",
        ),
        (12, 3): ("Sequencing info", 1, "Mapped reads, %"),
        (12, 4): ("Sequencing info", 1, "Chimeric DNA fragments, %"),
        (12, 5): ("Sequencing info", 1, "Insert size median, bp"),
        (12, 6): (
            "Sequencing info",
            1,
            "Unevenness of local genome coverage, x",
        ),
        (1, 1): "=SOC!A2",
        (2, 1): "=SOC!A3",
        (1, 3): "=SOC!A5",
        (2, 3): "=SOC!A6",
        (1, 5): "=SOC!A9",
        (15, 1): "QC alerts",
        (16, 1): "None",
        (15, 2): "Assessed purity",
        (15, 3): "SNV TMB",
    },
    "alignment_info": [
        (f"{col}{row}", {"horizontal": "center", "wrapText": True})
        for col in list("ABCDEFGH")
        for row in range(4, 6)
    ]
    + [
        (f"{col}{row}", {"horizontal": "center", "wrapText": True})
        for col in list("ABCDEFG")
        for row in range(7, 9)
    ]
    + [
        (f"{col}{row}", {"horizontal": "center", "wrapText": True})
        for col in list("ABCDEF")
        for row in range(10, 13)
    ]
    + [(f"{col}15", {"horizontal": "center"}) for col in list("ABC")],
    "to_bold": [f"{col}4" for col in list("ABCDEFGH")]
    + [f"{col}7" for col in list("ABCDEFG")]
    + [f"{col}10" for col in list("ABCDEF")]
    + [f"{col}15" for col in list("ABC")],
    "col_width": [
        ("A", 12),
        ("B", 12),
        ("C", 12),
        ("D", 12),
        ("E", 12),
        ("F", 12),
        ("G", 12),
        ("H", 12),
        ("I", 12),
        ("J", 12),
    ],
    "row_height": [(4, 30), (5, 30), (7, 30), (10, 30)],
    "cells_to_colour": [
        (
            f"{col}4",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for col in list("ABCDEFGH")
    ]
    + [
        (
            f"{col}7",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for col in list("ABCDEFG")
    ]
    + [
        (
            f"{col}10",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for col in list("ABCDEF")
    ],
    "borders": {
        "single_cells": [
            ("A15", LOWER_BORDER),
            ("B15", LOWER_BORDER),
            ("C15", LOWER_BORDER),
        ],
        "cell_rows": [
            ("A4:H4", THIN_BORDER),
            ("A5:H5", THIN_BORDER),
            ("A7:G7", THIN_BORDER),
            ("A8:G8", THIN_BORDER),
            ("A10:F10", THIN_BORDER),
            ("A11:F11", THIN_BORDER),
            ("A12:F12", THIN_BORDER),
        ],
    },
    "dropdowns": [
        {
            "cells": {
                ("A16",): (
                    '"None,'
                    "<30% tumour purity,"
                    "SNVs low VAF (<6%),"
                    "TINC (<5%),"
                    "TINC (>5%),"
                    "Tumour potentially degraded,"
                    "Tumour likely degraded,"
                    'Poor quality germline CNV calls"'
                ),
            },
            "title": "QC alerts",
        },
        {
            "cells": {
                ("B16",): ('"High (>70%),Medium (30-70%),Low (<30%)"'),
            },
            "title": "Assessed purity",
        },
        {
            "cells": {
                ("C16",): (
                    '"Not hypermutated (<10 mut/Mb),'
                    "Paed hypermutated (2-10 mut/Mb),"
                    "Hypermutated (>10 mut/Mb),"
                    'Ultra-hypermutaed (>100 mut/Mb)"'
                ),
            },
            "title": "SNV TMB",
        },
    ],
    "images": [
        {"cell": "E15", "img_index": 8, "size": (350, 500)},
        {"cell": "K15", "img_index": 10, "size": (350, 500)},
    ],
}
