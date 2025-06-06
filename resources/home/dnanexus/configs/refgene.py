from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc

THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)
LEFT_BORDER = Border(left=THIN)


CONFIG = {
    "cells_to_colour": [
        (
            f"{misc.convert_index_to_letters(i)}1",
            PatternFill(patternType="solid", start_color="dbeef4"),
        )
        for i in range(18 + 1)
    ],
    "borders": {
        "cell_rows": [
            ("B1:B1500", LEFT_BORDER),
            ("E1:E1500", LEFT_BORDER),
            ("H1:H1500", LEFT_BORDER),
            ("K1:K1500", LEFT_BORDER),
            ("N1:N1500", LEFT_BORDER),
            ("Q1:Q1500", LEFT_BORDER),
            ("T1:T1500", LEFT_BORDER),
        ],
    },
}


SHEETS2COLUMNS = {
    "somatic_db": {
        "Gene": "Gene",
        "Role in Cancer": "Comments",
        "Driver_SV": "COSMIC_Alteration",
        "Entities": "COSMIC_Entities",
    },
    "haem": {
        "Gene": "Gene",
        "Driver": "Haem_Alteration",
        "Entities": "Haem_Entities",
        "Comments": "Haem_Comments",
    },
    "paed": {
        "Gene": "Gene",
        "Driver": "Paed_Alteration",
        "Entities": "Paed_Entities",
        "Comments": "Paed_Comments",
    },
    "ovarian": {
        "Gene": "Gene",
        "Driver": "Ovarian_Alteration",
        "Entities": "Ovarian_Entities",
        "Comments": "Ovarian_Comments",
    },
    "sarc": {
        "Gene": "Gene",
        "Driver": "Sarcoma_Alteration",
        "Entities": "Sarcoma_Entites",
        "Comments": "Sarcoma_Comments",
    },
    "neuro": {
        "Gene": "Gene",
        "Driver": "Neuro_Alteration",
        "Entities": "Neuro_Entities",
        "Comments": "Neuro_Comments",
    },
}

RESCUE_COLUMNS = {"somatic_db": ["cosmic"]}


def add_dynamic_values(df: pd.DataFrame) -> dict:
    """Add dynamic values for the refgene sheet

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe containing the data for refgene

    Returns
    -------
    dict
        Dict containing data that needs to be merged to the CONFIG variable
    """

    sv_column_letter = misc.get_column_letter_using_column_name(df, "SNV")
    last_column_letter = misc.get_column_letter_using_column_name(df)

    config_with_dynamic_values = {
        "cells_to_write": {
            (1, i): column for i, column in enumerate(df.columns, 1)
        }
        | {
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(df), 1)
            for c_idx, value in enumerate(row, 1)
            # remove the col and row index from the writing
            if c_idx != 1 and r_idx != 1
        },
        "cells_to_colour": [
            (
                f"{misc.convert_index_to_letters(i)}1",
                PatternFill(patternType="solid", start_color="b686da"),
            )
            for i in range(
                misc.convert_letter_column_to_index(sv_column_letter),
                len(df.columns),
            )
        ],
        "to_bold": [
            f"{misc.convert_index_to_letters(i)}1"
            for i in range(
                misc.convert_letter_column_to_index(last_column_letter) + 1
            )
        ],
        "auto_filter": f"A:{last_column_letter}",
        "borders": {
            "cell_rows": [(f"A1:{last_column_letter}1", THIN_BORDER)],
        },
    }

    return config_with_dynamic_values
