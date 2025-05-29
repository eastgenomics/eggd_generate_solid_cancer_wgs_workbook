import string

from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc

THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)
LEFT_BORDER = Border(left=THIN)

CONFIG = {
    "cells_to_write": {
        (1, i): value
        for i, value in enumerate(
            [
                "Domain",
                "Gene",
                "GRCh38 coordinates",
                "Cyto",
                "RefSeq IDs",
                "Variant",
                "Predicted consequences",
                "Error flag",
                "Population germline allele frequency (GE | gnomAD)",
                "VAF",
                "LOH",
                "Alt allele/total read depth",
                "Gene mode of action",
                "Variant class",
                "TSG_NMD",
                "TSG_LOH",
                "Splice fs?",
                "SpliceAI",
                "REVEL",
                "OG_3' Ter",
                "Recurrence somatic database",
                "HS_Total",
                "HS_Mut",
                "HS_Tissue",
                "COSMIC Driver",
                "COSMIC Entities",
                "Paed Driver",
                "Paed Entities",
                "Sarc Driver",
                "Sarc Entities",
                "Neuro Driver",
                "Neuro Entities",
                "Ovary Driver",
                "Ovary Entities",
                "Haem Driver",
                "Haem Entities",
                "MTBP c.",
                "MTBP p.",
            ],
            1,
        )
    },
    "to_bold": [f"{misc.convert_index_to_letters(i)}1" for i in range(0, 38)],
    "col_width": [
        ("A", 5),
        ("B", 12),
        ("C", 20),
        ("D", 14),
        ("E", 20),
        ("F", 22),
        ("G", 22),
        ("K", 8),
        ("M", 18),
        ("N", 14),
    ]
    + [(f"{misc.convert_index_to_letters(i)}", 5) for i in range(21, 38)],
    "cells_to_colour": [
        # letters N to U
        (
            f"{string.ascii_uppercase[i]}1",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for i in range(13, 21)
    ]
    + [
        (
            f"{letter}1",
            PatternFill(patternType="solid", start_color="fdeada"),
        )
        for letter in ["V", "W", "X"]
    ]
    + [
        # letters Y to AJ
        (
            f"{misc.convert_index_to_letters(i)}1",
            PatternFill(patternType="solid", start_color="dbeef4"),
        )
        for i in range(24, 36)
    ]
    + [
        (
            f"{col}1",
            PatternFill(patternType="solid", start_color="dabcff"),
        )
        for col in ["AK", "AL"]
    ],
    "borders": {
        "cell_rows": [
            ("A1:AL1", THIN_BORDER),
        ],
    },
    "alignment_info": [
        (
            f"{misc.convert_index_to_letters(i)}1",
            {
                "horizontal": "left",
                "vertical": "bottom",
                "wrapText": True,
                "text_rotation": 90,
            },
        )
        for i in range(0, 38)
    ],
    "row_height": [(1, 80)],
    "auto_filter": "A:AL",
    "freeze_panes": "G1",
}


def add_dynamic_values(data: pd.DataFrame) -> dict:
    """Add dynamic values for the SNV sheet

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe containing the data for somatic variants and appropriate
        additional data from inputs

    Returns
    -------
    dict
        Dict containing data that needs to be merged to the CONFIG variable
    """

    nb_somatic_variants = data.shape[0]

    config_with_dynamic_values = {
        "cells_to_write": {
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(data), 1)
            for c_idx, value in enumerate(row, 1)
            # remove the col and row index from the writing?
            if c_idx != 1 and r_idx != 1
        },
        "dropdowns": [
            {
                "cells": {
                    (f"N{i}" for i in range(2, nb_somatic_variants + 2)): (
                        '"Oncogenic, Likely oncogenic,'
                        "Uncertain, Likely passenger,"
                        'Likely artefact"'
                    ),
                },
                "title": "Variant class",
            },
        ],
        "data_bar": f"J2:J{nb_somatic_variants + 1}",
        "borders": {
            "cell_rows": [
                (f"Y1:Y{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AA1:AA{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AC1:AC{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AE1:AE{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AG1:AG{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AI1:AI{nb_somatic_variants+1}", LEFT_BORDER),
                (f"AK1:AK{nb_somatic_variants+1}", LEFT_BORDER),
            ],
        },
    }

    return config_with_dynamic_values
