import re

from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
import pandas as pd

from utils import misc

# prepare formatting
THIN = Side(border_style="thin", color="000000")
THICK = Side(border_style="thick", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)

CONFIG = {
    "cells_to_write": {
        # hardcoded table headers and other elements
        (1, 1): "=SOC!A2",
        (2, 1): "=SOC!A3",
        (1, 3): "=SOC!A5",
        (2, 3): "=SOC!A6",
        (1, 5): "=SOC!A9",
        (23, 1): "Somatic SNV",
        (24, 1): "Gene",
        (24, 2): "GRCh38 Coordinates",
        (24, 3): "Variant",
        (24, 4): "Consequence",
        (24, 5): "VAF",
        (24, 6): "Variant Class",
        (24, 7): "Actionability",
        (24, 8): "Comments",
        (24, 9): "TOC",
        (35, 1): "Somatic CNV_SV",
        (36, 1): "Gene/Locus",
        (36, 2): "GRCh38 Coordinates",
        (36, 3): "Cytological Bands",
        (36, 4): "Variant Type",
        (36, 5): "Consequence",
        (36, 6): "Variant Class",
        (36, 7): "Actionability",
        (36, 8): "Comments",
        (36, 9): "TOC",
        (53, 1): "Germline SNV",
        (54, 1): "Gene",
        (54, 2): "GRCh38 Coordinates",
        (54, 3): "Variant",
        (54, 4): "Consequence",
        (54, 5): "Zygosity",
        (54, 6): "Tumour VAF",
        (54, 7): "Variant Class",
        (54, 8): "Actionability",
        (54, 9): "Comments",
        (54, 10): "TOC",
        (60, 1): "Germline CNV",
        (61, 1): "Gene",
        (61, 2): "GRCh38 Coordinates",
        (61, 3): "Variant",
        (61, 4): "Consequence",
        (61, 5): "Zygosity",
        (61, 6): "Variant Class",
        (61, 7): "Actionability",
        (61, 8): "Comments",
        (61, 9): "TOC",
        (67, 1): "Somatic_SNV",
        (79, 1): "Somatic_CNV_GAIN",
        (87, 1): "Somatic_CNV_LOSS",
        (95, 1): "Somatic_SV",
        (103, 1): "Germline_SNV",
        (110, 1): "Germline_CNV",
        # summary to be pasted
        (3, 8): "TMB (Mut/Mb)",
        (3, 9): "=QC!G8",
        (4, 8): "Pertinent Signatures",
        (4, 9): "=Signatures!C36",
        (5, 8): "Somatic Chr aberrations",
        (5, 9): "=Plot!A35",
        (6, 8): "Somatic SNV/indel",
        (6, 9): '=_xlfn.TEXTJOIN(", ",TRUE(),A25:A33)',
        (7, 8): "Somatic CNV Gain",
        (7, 9): '=_xlfn.TEXTJOIN(", ",TRUE(),A37:A41)',
        (8, 8): "Somatic CNV Loss",
        (8, 9): '=_xlfn.TEXTJOIN(", ",TRUE(),A42:A46)',
        (9, 8): "Somatic SV",
        (10, 8): "Somatic VUS",
        (11, 8): "Germline",
        (12, 8): "GTAB date",
        (13, 8): "SOC genes reported",
        (13, 9): "=SOC!A13",
        (14, 8): "WGS novel genes",
        (15, 8): "Histological diagnosis",
        (15, 9): "=SOC!A9",
        (16, 8): "QC alerts",
        (16, 9): "=QC!A16",
        (17, 8): "Genotype = histo dx.",
        (18, 8): "Actionable genes",
        (19, 8): "Referral to ClinGen",
        (20, 8): "GTAB advice",
        (21, 8): "Forwarding recipients",
        # outcode codes
        (3, 11): "Testing outcome codes (TOC)",
        (4, 11): "411",
        (4, 12): "Variant contributes to dx",
        (5, 11): "412",
        (5, 12): "Variant contributes to alternative dx",
        (6, 11): "413",
        (
            6,
            12,
        ): "Variant reduces likelihood but does not exclude differential dx",
        (7, 11): "421",
        (
            7,
            12,
        ): "Variant informs targeted treatment or prognostic/actionable information",
        (8, 11): "422",
        (
            8,
            12,
        ): "Wild-type result, absence of variant means targeted treatment not available",
        (9, 11): "423",
        (
            9,
            12,
        ): "Wild-type result, absence of variant means targeted treatment is available or where Prognostic/actionable information is provided",
        (10, 11): "971",
        (10, 12): "Failure",
        (11, 11): "961",
        (11, 12): "Incidental finding",
        (12, 11): "991",
        (12, 12): "Other (not listed)",
        (13, 11): "992",
        (
            13,
            12,
        ): "Caveated result (e.g. no actionable variant, but low tumour purity so could be false negative)",
        (17, 11): "Lab comments",
    }
    ####
    # somatic snv gene lookup
    | {
        (row, 1): f'=SUBSTITUTE(B{row+44},";",CHAR(10))'
        for row in range(25, 34)
    }
    # somatic snv coordinates
    | {
        (row, 2): f'=SUBSTITUTE(C{row+44},";",CHAR(10))'
        for row in range(25, 34)
    }
    # somatic snv variant
    | {
        (row, 3): f'=SUBSTITUTE(F{row+44},";",CHAR(10))'
        for row in range(25, 34)
    }
    # somatic snv consequences
    | {(row, 4): f"=G{row+44}" for row in range(25, 34)}
    # somatic snv VAF
    | {
        (row, 5): f"=CONCATENATE(J{row+44},CHAR(10),K{row+44})"
        for row in range(25, 34)
    }
    # somatic snv variant class
    | {(row, 6): f"=N{row+44}" for row in range(25, 34)}
    ####
    # somatic cnv gain lookup
    | {
        (row, 1): f'=SUBSTITUTE(B{row+44},";",CHAR(10))'
        for row in range(37, 42)
    }
    # somatic cnv gain coordinates
    | {
        (row, 2): f'=SUBSTITUTE(E{row+44},";",CHAR(10))'
        for row in range(37, 42)
    }
    # somatic cnv gain cytological bands
    | {
        (row, 3): f"=CONCATENATE(I{row+44},CHAR(10),J{row+44})"
        for row in range(37, 42)
    }
    # somatic cnv gain variant type
    | {
        (row, 4): f'=CONCATENATE(F{row+44}," (",G{row+44},")")'
        for row in range(37, 42)
    }
    # somatic cnv gain variant class
    | {(row, 6): f"=L{row+44}" for row in range(37, 42)}
    ####
    # somatic cnv loss lookup
    | {
        (row, 1): f'=SUBSTITUTE(B{row+47},";",CHAR(10))'
        for row in range(42, 47)
    }
    # somatic cnv loss coordinates
    | {
        (row, 2): f'=SUBSTITUTE(E{row+47},";",CHAR(10))'
        for row in range(42, 47)
    }
    # somatic cnv loss cytological bands
    | {
        (row, 3): f"=CONCATENATE(I{row+47},CHAR(10),J{row+47})"
        for row in range(42, 47)
    }
    # somatic cnv loss variant type
    | {
        (row, 4): f'=CONCATENATE(F{row+47}," (",G{row+47},")")'
        for row in range(42, 47)
    }
    # somatic cnv loss variant class
    | {(row, 6): f"=L{row+47}" for row in range(42, 47)}
    ####
    # somatic fusion gene lookup
    | {
        (row, 1): f'=SUBSTITUTE(B{row+50},";",CHAR(10))'
        for row in range(47, 52)
    }
    # somatic fusion coordinates
    | {
        (row, 2): f'=SUBSTITUTE(E{row+50},";",CHAR(10))'
        for row in range(47, 52)
    }
    | {
        (
            row,
            4,
        ): f'=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(F{row+50},"BND","Translocation"),"INV","Inversion"),"DEL","Deletion"),"DUP","Tandem duplication")'
        for row in range(47, 52)
    }
    ####
    # germline snv gene lookup
    | {(row, 1): f"=A{row+50}" for row in range(55, 59)}
    # germline snv coordinates lookup
    | {(row, 2): f"=B{row+50}" for row in range(55, 59)}
    # germline snv variant lookup
    | {(row, 3): f"=C{row+50}" for row in range(55, 59)}
    # germline snv consequence lookup
    | {(row, 4): f"=D{row+50}" for row in range(55, 59)}
    # germline snv tumour vaf lookup
    | {(row, 6): f"=I{row+50}" for row in range(55, 59)},
    ####
    "to_bold": [
        # table names to be bolded
        "A1",
        "A23",
        "A35",
        "A53",
        "A60",
        "A67",
        "A79",
        "A87",
        "A95",
        "A103",
        "A110",
    ]
    # table headers to be bolded
    + [f"{col}24" for col in list("ABCDEFGHI")]
    + [f"{col}36" for col in list("ABCDEFGHI")]
    + [f"{col}54" for col in list("ABCDEFGHIJ")]
    + [f"{col}61" for col in list("ABCDEFGHI")]
    + [f"H{row}" for row in range(3, 22)]
    + ["K3", "K17"],
    "col_width": [
        ("A", 10),
        ("B", 20),
        ("C", 16),
        ("D", 20),
        ("E", 15),
        ("F", 15),
        ("G", 24),
        ("H", 24),
        ("I", 24),
    ],
    "cells_to_colour": [
        (
            f"{column}{row}",
            PatternFill(patternType="solid", start_color="F2F2F2"),
        )
        for row in [24, 36, 54, 61]
        for column in list("ABCDEFGHI")
    ]
    + [("J54", PatternFill(patternType="solid", start_color="F2F2F2"))]
    + [
        (f"H{row}", PatternFill(patternType="solid", start_color="dce6f2"))
        for row in range(3, 12)
    ]
    + [
        (f"H{row}", PatternFill(patternType="solid", start_color="fdeada"))
        for row in range(12, 22)
    ]
    + [
        (f"{col}3", PatternFill(patternType="solid", start_color="fdeada"))
        for col in list("KLM")
    ]
    + [
        (f"{col}17", PatternFill(patternType="solid", start_color="fdeada"))
        for col in list("KLM")
    ],
    "borders": {
        "cell_rows": [(f"A{row}:I{row}", THIN_BORDER) for row in range(24, 34)]
        + [(f"A{row}:I{row}", THIN_BORDER) for row in range(36, 52)]
        + [(f"A{row}:J{row}", THIN_BORDER) for row in range(54, 59)]
        + [(f"A{row}:I{row}", THIN_BORDER) for row in range(61, 65)]
        + [("H11:I11", LOWER_BORDER)]
    },
    "alignment_info": [
        (
            f"{col}{row}",
            {
                "wrapText": True,
                "horizontal": "center",
                "vertical": "center",
            },
        )
        for col in list("ABCDEFGHI")
        for row in range(24, 34)
    ]
    + [
        (
            f"{col}{row}",
            {
                "wrapText": True,
                "horizontal": "center",
                "vertical": "center",
            },
        )
        for col in list("ABCDEFGHI")
        for row in range(36, 52)
    ]
    + [
        (
            f"{col}{row}",
            {
                "wrapText": True,
                "horizontal": "center",
                "vertical": "center",
            },
        )
        for col in list("ABCDEFGHIJ")
        for row in range(54, 59)
    ]
    + [
        (
            f"{col}{row}",
            {
                "wrapText": True,
                "horizontal": "center",
                "vertical": "center",
            },
        )
        for col in list("ABCDEFGHI")
        for row in range(61, 66)
    ]
    + [
        (
            f"{col}{row}",
            {
                "wrapText": True,
                "horizontal": "center",
                "vertical": "center",
            },
        )
        for col in list("KLM")
        for row in [3, 17]
    ],
    "row_height": [
        (row, 30)
        for start, end in [(25, 34), (37, 52), (54, 59), (61, 65)]
        for row in range(start, end)
    ],
    "dropdowns": [
        {
            "cells": {
                ("I17",): ('"Yes,' "No," '-"'),
            },
            "title": "Genotype = histo dx.",
        },
        {
            "cells": {
                ("I19",): ('"Yes,' "No," 'Previously known"'),
            },
            "title": "Referral to ClinGen",
        },
        {
            "cells": {
                (
                    f"G{row}"
                    for start, end in [(25, 34), (37, 52)]
                    for row in range(start, end)
                ): (
                    '"Predicts therapeutic response,'
                    "Prognostic,"
                    "Defines diagnosis group,"
                    "Eligibility for trial,"
                    'Other"'
                ),
            },
            "title": "Actionability",
        },
        {
            "cells": {
                (
                    f"G{row}"
                    for start, end in [(55, 60)]
                    for row in range(start, end)
                ): ('"Pathogenic,Likely pathogenic,Uncertain"'),
            },
            "title": "Variant class germline",
        },
        {
            "cells": {
                (
                    f"H{row}"
                    for start, end in [(55, 60)]
                    for row in range(start, end)
                ): (
                    '"Predicts therapeutic response,'
                    "Prognostic,"
                    "Defines diagnosis group,"
                    "Eligibility for trial,"
                    'Other"'
                ),
            },
            "title": "Actionability Germline",
        },
    ],
    "to_merge": [
        {
            "start_row": 3,
            "end_row": 3,
            "start_column": 11,
            "end_column": 13,
        },
        {
            "start_row": 17,
            "end_row": 17,
            "start_column": 11,
            "end_column": 13,
        },
    ],
}


def add_dynamic_values(
    SV_df: pd.DataFrame,
    fusion_count: int,
    nb_germline_genes: int,
    SNV_df_columns: list = None,
    gain_df_columns: list = None,
    loss_df_columns: list = None,
    SV_df_columns: list = None,
    germline_df_columns: list = None,
) -> dict:
    """Add dynamic values for the Summary sheet

    Parameters
    ----------
    SV_df : pd.DataFrame
        Dataframe containing the data for SV fusion variants and appropriate
        additional data from inputs
    fusion_count : int
        Integer for the maximum number of fusion for a variant
    SNV_df_columns : list
        List of columns for the SNV dataframe
    gain_df_columns : list
        List of columns for the gain dataframe
    loss_df_columns : list
        List of columns for the loss dataframe
    SV_df_columns : list
        List of columns for the SV dataframe
    germline_df_columns : list
        List of columns for the germline dataframe

    Returns
    -------
    dict
        Dict containing data that needs to be merged to the CONFIG variable
    """

    variant_class_column_letter = misc.get_column_letter_using_column_name(
        SV_df, "Variant class"
    )
    actionability_column_letter = misc.get_column_letter_using_column_name(
        SV_df, "Actionability"
    )
    comments_column_letter = misc.get_column_letter_using_column_name(
        SV_df, "Comments"
    )

    sv_pair = [
        (1, "C"),
        (2, "D"),
        (3, "E"),
        (4, "F"),
        (5, "G"),
        (6, variant_class_column_letter),
        (7, actionability_column_letter),
        (8, comments_column_letter),
    ]

    if fusion_count == 0:
        sv_pair.pop(4)

    all_df_columns = [
        {
            (row, col_index): col_name
            for col_index, col_name in enumerate(df_columns, 1)
        }
        for df_columns, row in [
            (SNV_df_columns, 68),
            (gain_df_columns, 80),
            (loss_df_columns, 88),
            (SV_df_columns, 96),
            (germline_df_columns, 104),
        ]
        if df_columns is not None
    ]

    # the position of the variant class and cyto columns is going to be dynamic
    # depending on the number of fusion elements. This attempts to get the
    # positions of those columns
    *cytos_column_index, variant_class_column_index = sorted(
        [
            index
            for index, col in enumerate(SV_df_columns)
            for col_to_find in ["Variant class", "Cyto"]
            if re.match(col_to_find, col)
        ]
    )

    config_with_dynamic_values = {
        "cells_to_write": {
            key: value
            for data_dict in all_df_columns
            for key, value in data_dict.items()
        }
        | {
            # 7 is the nb of rows before and after the germline table in the
            # germline sheet
            (11, 9): f"=Germline!A{nb_germline_genes + 7}",
        }
        # dynamic way to concatenate as many cyto bands as possible, i'm sorry
        | {
            (row, 3): "=CONCATENATE("
            + ",CHAR(10),".join(
                [
                    f"{misc.convert_index_to_letters(cyto)}{row+50}"
                    for cyto in cytos_column_index
                ]
            )
            + ")"
            for row in range(47, 52)
        }
        | {
            (
                row,
                6,
            ): f"={misc.convert_index_to_letters(variant_class_column_index)}{row+50}"
            for row in range(47, 52)
        },
    }

    return config_with_dynamic_values
