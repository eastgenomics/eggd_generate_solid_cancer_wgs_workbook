import re

from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc

# prepare formatting
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT_BORDER = Border(left=THIN)


CONFIG = {
    "col_width": [
        ("A", 12),
        ("B", 18),
        ("C", 22),
        ("D", 22),
        ("E", 20),
    ],
    "expected_columns": [
        "Event domain",
        "Gene",
        "RefSeq IDs",
        "Impacted transcript region",
        "GRCh38 coordinates",
        "Size",
        (
            "Population germline allele frequency (GESG | GECG for somatic "
            "SVs or AF | AUC for germline CNVs)"
        ),
        "Paired reads",
        "Split reads",
        "Gene mode of action",
        "Variant class",
        "Comments",
        "OG_Fusion",
        "OG_IntDup",
        "OG_IntDel",
        "Disruptive",
    ],
    "alternative_columns": [
        [
            (
                "Population germline allele frequency (GESG | GECG for "
                "somatic SVs or AF | AUC for germline CNVs)"
            ),
            (
                "Population germline allele frequency (AF | AUC for germline "
                "CNVs)"
            ),
        ]
    ],
    "row_height": [(1, 120)],
}


def add_dynamic_values(data: pd.DataFrame, alternative_columns: dict) -> dict:
    """Add dynamic values for the SV sheet

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe containing the data for SV fusion variants and appropriate
        additional data from inputs
    alternative_columns: dict
        Dict containing the expected headers and the applied alternatives

    Returns
    -------
    dict
        Dict containing data that needs to be merged to the CONFIG variable
    """

    if data is None:
        return {}

    nb_structural_variants = data.shape[0]

    column_letters = []

    for column_name in ["Size", "Variant class", "Gene mode of action"]:
        if column_name in alternative_columns:
            column_name = alternative_columns[column_name]

        column_letters.append(
            misc.get_column_letter_using_column_name(data, column_name)
        )

    last_column_letter = misc.get_column_letter_using_column_name(data)
    variant_class_column_index = misc.convert_letter_column_to_index(
        column_letters[1]
    )

    first_letter_lookup_groups = variant_class_column_index + 6

    lookup_start, last_column_index = (
        first_letter_lookup_groups,
        misc.convert_letter_column_to_index(last_column_letter),
    )

    total_number_genes = last_column_index - lookup_start + 1

    # there are 13 look up groups
    if total_number_genes % 13:
        raise ValueError(
            (
                "Uneven number of genes per lookup group: "
                f"{total_number_genes} / 13 = {total_number_genes/13} per "
                "group"
            )
        )

    number_genes = total_number_genes // 13

    number_genes = int(number_genes)
    lookup_end = last_column_index - number_genes

    border_cells = []

    # build the info for borders in the lookup groups
    # +2 in order to add a left border between the lookup groups and the gene
    # columns
    # number_genes*2 to add borders at start and end of each lookup group
    # rather than between drivers and entities of each lookup
    for i, index in enumerate(range(lookup_start, last_column_index + 2)):
        if i % (number_genes*2) == 0:
            col_letter = misc.convert_index_to_letters(index)
            border_cells.append(
                (
                    f"{col_letter}2:{col_letter}{nb_structural_variants+1}",
                    LEFT_BORDER,
                )
            )

    config_with_dynamic_values = {
        "cells_to_write": {
            (1, i): column for i, column in enumerate(data.columns, 1)
        }
        | {
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(data), 1)
            for c_idx, value in enumerate(row, 1)
            # remove the col and row index from the writing
            if c_idx != 1 and r_idx != 1
        },
        "cells_to_colour": [
            (
                f"{misc.convert_index_to_letters(i)}1",
                PatternFill(patternType="solid", start_color="F2F2F2"),
            )
            for i in range(
                variant_class_column_index, variant_class_column_index + 6
            )
        ]
        + [
            (
                f"{misc.convert_index_to_letters(i)}1",
                PatternFill(patternType="solid", start_color="dbeef4"),
            )
            for i in range(lookup_start, lookup_end + 1)
        ]
        + [
            (
                f"{misc.convert_index_to_letters(i)}1",
                PatternFill(patternType="solid", start_color="e6e0ec"),
            )
            for i in range(lookup_end + 1, last_column_index + 1)
        ],
        "to_bold": [
            f"{misc.convert_index_to_letters(i)}1"
            for i in range(last_column_index + 1)
        ],
        # define width for the empty columns that scientists fill in
        "col_width": [
            (misc.convert_index_to_letters(i), 6)
            for i in range(
                variant_class_column_index + 1, variant_class_column_index + 5
            )
        ]
        # define width for Gene mode of action
        + [(column_letters[2], 22)]
        # define width for the Fusion columns
        + [
            (misc.convert_index_to_letters(i), 20)
            for i, column in enumerate(data.columns)
            if re.match(r"Fusion_[0-9]+", column)
        ],
        "borders": {
            "cell_rows": [(f"A1:{last_column_letter}1", THIN_BORDER)]
            + border_cells
        },
        "alignment_info": [
            (
                f"{misc.convert_index_to_letters(i)}1",
                {
                    "horizontal": "left",
                    "vertical": "bottom",
                    "wrapText": True,
                    "text_rotation": 90,
                },
            )
            for i in range(0, lookup_end + 1)
        ],
        "freeze_panes": f"{column_letters[0]}1",
        "dropdowns": [
            {
                "cells": {
                    (
                        f"{column_letters[1]}{i}"
                        for i in range(2, nb_structural_variants + 2)
                    ): (
                        '"Oncogenic, Likely oncogenic,'
                        "Uncertain, Likely passenger,"
                        'Likely artefact"'
                    ),
                },
                "title": "Variant class",
            },
        ],
        "auto_filter": f"F:{last_column_letter}",
    }

    return config_with_dynamic_values
