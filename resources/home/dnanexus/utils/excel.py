import re

from bs4 import BeautifulSoup
import openpyxl
from openpyxl import drawing
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from PIL import Image
import vcfpy

from configs.tables import get_table_value_in_html_table
from utils import misc, vcf

pd.options.mode.chained_assignment = None


def open_file(file: str, file_type: str) -> pd.DataFrame:
    """Read in CSV or XLS files using pandas

    Parameters
    ----------
    file : str
        File path
    file_type : str
        File type with the file path

    Returns
    -------
    pd.DataFrame
        Dataframe created by pandas
    """

    if file_type == "csv":
        df = pd.read_csv(file)
    elif file_type == "xls":
        df = pd.read_excel(file, sheet_name=None)

    # convert the clinvar id column as a string and remove the trailing .0 that
    # the automatic conversion that pandas applies added
    if df is pd.DataFrame and "ClinVar ID" in df.columns:
        df["ClinVar ID"] = df["ClinVar ID"].astype(str)
        df["ClinVar ID"] = df["ClinVar ID"].str.removesuffix(".0")

    return df


def process_reported_variants_germline(
    df: pd.DataFrame, clinvar_resource: vcfpy.Reader
) -> pd.DataFrame:
    """Process the data from the reported variants excel file

    Parameters
    ----------
    df : pd.DataFrame
        Dataframe from parsing the reported variants excel file
    clinvar_resource : vcfpy.Reader
        vcfpy.Reader object from the Clinvar resource

    Returns
    -------
    pd.DataFrame
        Dataframe containing clinical significance info for germline variants
    """

    df = df[df["Origin"].str.lower() == "germline"]

    if df.empty:
        return None

    df.reset_index(drop=True, inplace=True)

    clinvar_ids_to_find = [
        value for value in df.loc[:, "ClinVar ID"].to_numpy()
    ]
    clinvar_info = vcf.find_clinvar_info(
        clinvar_resource, *clinvar_ids_to_find
    )

    # add the clinvar info by merging the clinvar dataframe
    df = df.merge(clinvar_info, on="ClinVar ID", how="left")

    # split the col to get gnomAD
    df[["GE", "gnomAD"]] = df[
        "Population germline allele frequency (GE | gnomAD)"
    ].str.split("|", expand=True)

    df.drop(
        ["GE", "Population germline allele frequency (GE | gnomAD)"],
        axis=1,
        inplace=True,
    )
    df.loc[:, "Variant Class"] = ""
    df.loc[:, "Actionability"] = ""
    df = df[
        [
            "Gene",
            "GRCh38 coordinates;ref/alt allele",
            "CDS change and protein change",
            "Predicted consequences",
            "Genotype",
            "Variant Class",
            "Actionability",
            "Gene mode of action",
            "clnsigconf",
            "gnomAD",
        ]
    ]

    df.fillna("", inplace=True)

    return df


def process_reported_variants_somatic(
    df: pd.DataFrame, refgene_dfs: dict, hotspots_df: pd.DataFrame
) -> pd.DataFrame:
    """Get the somatic variants and format the data for them

    Parameters
    ----------
    df : pd.DataFrame
        Dataframe from parsing the reported variants excel file
    refgene_dfs : dict
        Dict of dataframes from the refgene excel
    hotspots_df : pd.DataFrame
        Dataframe from parsing the hotspots excel file

    Returns
    -------
    pd.DataFrame
        Dataframe with additional formatting for c. and p. annotation
    """

    # select only somatic rows
    df = df[df["Origin"].str.lower().str.contains("somatic")]
    df.reset_index(drop=True, inplace=True)
    df[["c_dot", "p_dot"]] = df["CDS change and protein change"].str.split(
        r"(?=;p)", n=1, expand=True
    )
    df["p_dot"] = df["p_dot"].str.slice(1)

    df["MTBP c."] = df["Gene"] + ":" + df["c_dot"]
    df["MTBP p."] = df["Gene"] + ":" + df["p_dot"]
    df.fillna({"MTBP p.": ""}, inplace=True)

    # convert string like: NRAS:p.Gln61Arg to NRAS:p.Gln61 for lookup in the
    # hotspots excel
    df["HS p."] = df["MTBP p."].apply(
        lambda x: (
            x[: re.search(r":p.[A-Za-z]+[0-9]+", x).end()]
            if re.search(r":p.[A-Za-z]+[0-9]+", x)
            else x
        )
    )

    # populate the somatic variant dataframe with data from the refgene excel
    # file
    lookup_refgene = (
        ("COSMIC", "Gene", refgene_dfs["cosmic"], "Gene", "Entities"),
        ("Paed", "Gene", refgene_dfs["paed"], "Gene", "Driver"),
        ("Sarc", "Gene", refgene_dfs["sarc"], "Gene", "Driver"),
        ("Neuro", "Gene", refgene_dfs["neuro"], "Gene", "Driver"),
        ("Ovary", "Gene", refgene_dfs["ovarian"], "Gene", "Driver"),
        ("Haem", "Gene", refgene_dfs["haem"], "Gene", "Driver"),
        ("HS_Sample", "HS p.", hotspots_df, "HS_PROTEIN_ID", "HS_Samples"),
        (
            "HS_Tumour",
            "HS p.",
            hotspots_df,
            "HS_PROTEIN_ID",
            "HS_Tumor Type Composition",
        ),
    )

    for (
        new_column,
        col_to_map,
        reference_df,
        col_to_index,
        col_to_look_up,
    ) in lookup_refgene:
        df[new_column] = df[col_to_map].map(
            reference_df.set_index(col_to_index)[col_to_look_up]
        )
        df[new_column] = df[new_column].fillna("-")

    df.loc[:, "Error flag"] = ""

    df["con_count"] = df["Predicted consequences"].str.count(r"\;")

    if df["con_count"].max() > 0:
        df[["Predicted consequences", "Error flag"]] = df[
            "Predicted consequences"
        ].str.split(";", expand=True)

    df.loc[:, "LOH"] = ""

    df["VAF"] = df["VAF"].astype("str")
    df["VAF_count"] = df["VAF"].str.count(r"\;")

    if df["VAF_count"].max() > 0:
        df[["VAF", "LOH"]] = df["VAF"].str.split(";", expand=True)

    df.loc[:, "Variant class"] = ""
    df.loc[:, "Actionability"] = ""
    df.loc[:, "Comments"] = ""
    df = df[
        [
            "Domain",
            "Gene",
            "GRCh38 coordinates;ref/alt allele",
            "CDS change and protein change",
            "Predicted consequences",
            "VAF",
            "LOH",
            "Error flag",
            "Alt allele/total read depth",
            "Gene mode of action",
            "Variant class",
            "Actionability",
            "Comments",
            "COSMIC",
            "Paed",
            "Sarc",
            "Neuro",
            "Ovary",
            "Haem",
            "HS_Sample",
            "HS_Tumour",
            "MTBP c.",
            "MTBP p.",
        ]
    ]
    df.rename(
        columns={
            "GRCh38 coordinates;ref/alt allele": "GRCh38 coordinates",
            "CDS change and protein change": "Variant",
        },
        inplace=True,
    )
    df.sort_values(["Domain", "VAF"], ascending=[True, False], inplace=True)
    df = df.replace([None], [""], regex=True)
    df["VAF"] = df["VAF"].astype(float)

    return df


def process_reported_SV(df: pd.DataFrame, refgene_dfs: dict) -> tuple:
    """Process the reported structural variants excel

    Parameters
    ----------
    df : pd.DataFrame
        Dataframe containing data from the structural variants excel
    refgene_dfs : dict
        Dict of dataframes from the refgene excel

    Returns
    -------
    tuple
        Tuple of the dataframes for Gain and Loss structural variants
    """

    df_loss = df[df["Type"].str.lower().str.contains("loss|loh")]
    df_loss.reset_index(drop=True, inplace=True)

    df_gain = df[df["Type"].str.lower().str.contains("gain")]
    df_gain.reset_index(drop=True, inplace=True)

    # populate the structural variant dataframe with data from the refgene
    # excel file
    lookup_refgene = (
        ("COSMIC", "Gene", refgene_dfs["cosmic"], "Gene", "Entities"),
        ("Paed", "Gene", refgene_dfs["paed"], "Gene", "Driver"),
        ("Sarc", "Gene", refgene_dfs["sarc"], "Gene", "Driver"),
        ("Neuro", "Gene", refgene_dfs["neuro"], "Gene", "Driver"),
        ("Ovary", "Gene", refgene_dfs["ovarian"], "Gene", "Driver"),
        ("Haem", "Gene", refgene_dfs["haem"], "Gene", "Driver"),
    )

    for sv_df in [df_loss, df_gain]:
        for (
            new_column,
            col_to_map,
            reference_df,
            col_to_index,
            col_to_look_up,
        ) in lookup_refgene:
            sv_df[new_column] = sv_df[col_to_map].map(
                reference_df.set_index(col_to_index)[col_to_look_up]
            )
            sv_df[new_column] = sv_df[new_column].fillna("-")

        sv_df.loc[:, "Variant class"] = ""
        sv_df.loc[:, "Actionability"] = ""
        sv_df.loc[:, "Comments"] = ""

        sv_df[["Type", "Copy Number"]] = sv_df.Type.str.split(
            r"\(|\)", expand=True
        ).iloc[:, [0, 1]]
        sv_df["Copy Number"] = sv_df["Copy Number"].astype(int)
        sv_df["Size"] = sv_df.apply(
            lambda x: "{:,.0f}".format(x["Size"]), axis=1
        )

        if list(sv_df["Type"].unique()) == ["GAIN"]:
            sv_df.sort_values(
                ["Event domain", "Copy Number"],
                ascending=[True, False],
                inplace=True,
            )
        else:
            sv_df.sort_values(
                ["Event domain", "Copy Number"],
                ascending=[True, True],
                inplace=True,
            )

    selected_col = [
        "Event domain",
        "Impacted transcript region",
        "Gene",
        "GRCh38 coordinates",
        "Chromosomal bands",
        "Type",
        "Copy Number",
        "Size",
        "Gene mode of action",
        "Variant class",
        "Actionability",
        "Comments",
        "COSMIC",
        "Paed",
        "Sarc",
        "Neuro",
        "Ovary",
        "Haem",
    ]
    df_loss = df_loss[selected_col]
    df_gain = df_gain[selected_col]

    return df_gain, df_loss


def process_refgene(dfs: dict) -> dict:
    """Process the refgene group excel by replacing the NA by * in select
    columns

    Parameters
    ----------
    dfs : dict
        Dict of dataframes corresponding to the data in the sheets in the
        refgene group excel

    Returns
    -------
    dict
        Dict of processed dataframes
    """

    for df in [
        dfs["cosmic"],
        dfs["paed"],
        dfs["sarc"],
        dfs["neuro"],
        dfs["ovarian"],
        dfs["haem"],
    ]:
        if "Entities" in list(df.columns):
            df["Entities"].astype(str)
            df.fillna({"Entities": "*"}, inplace=True)
        if "Driver" in list(df.columns):
            df["Driver"].astype(str)
            df.fillna({"Driver": "*"}, inplace=True)

    return dfs


def write_sheet(
    excel_writer: pd.ExcelWriter,
    sheet_name: str,
    html_tables: list = None,
    html_images: list = None,
    soup: BeautifulSoup = None,
    dynamic_data: dict = None,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Using a config file, write in the appropriate data

    Parameters
    ----------
    excel_writer : pd.ExcelWriter
        ExcelWriter object
    sheet_name : str
        Name of the sheet used to match the config
    html_tables : list, optional
        List of tables extracted from the HTML
    html_images : list, optional
        List of images extracted from the HTML
    soup : BeautifulSoup, optional
        BeautifulSoup object for the HTML file
    dynamic_data: dict, optional
        Dict of data for dynamic filling in the sheet

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        Worksheet object
    """

    sheet = excel_writer.book.create_sheet(sheet_name)

    type_config = misc.select_config(sheet_name)
    assert type_config, "Config file couldn't be imported"

    if dynamic_data:
        sheet_config = misc.merge_dicts(
            type_config.CONFIG, dynamic_data, sheet_name
        )
    else:
        sheet_config = type_config.CONFIG

    if sheet_config.get("cells_to_write"):
        write_cell_content(
            sheet, sheet_config["cells_to_write"], html_tables, soup
        )

    if sheet_config.get("to_merge"):
        # merge columns that have longer text
        sheet.merge_cells(**sheet_config["to_merge"])

    if sheet_config.get("to_align"):
        align_cells(sheet, sheet_config["to_align"])

    if sheet_config.get("to_bold"):
        bold_cells(sheet, sheet_config["to_bold"])

    if sheet_config.get("col_width"):
        set_col_width(sheet, sheet_config["col_width"])

    if sheet_config.get("cells_to_colour"):
        color_cells(sheet, sheet_config["cells_to_colour"])

    if sheet_config.get("borders"):
        draw_borders(sheet, sheet_config["borders"])

    if sheet_config.get("dropdowns"):
        generate_dropdowns(sheet, sheet_config["dropdowns"])

    if sheet_config.get("images"):
        insert_images(sheet, sheet_config["images"], html_images)

    if sheet_config.get("auto_filter"):
        filters = sheet.auto_filter
        filters.ref = sheet_config["auto_filter"]

    if sheet_config.get("freeze_panes"):
        sheet.freeze_panes = sheet[sheet_config["freeze_panes"]]

    if sheet_config.get("data_bar"):
        add_databar_rule(sheet, sheet_config["data_bar"])

    return sheet


def write_cell_content(
    sheet: Worksheet, config_data: dict, html_tables: list, soup: BeautifulSoup
):
    """Write the tables from the config

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to write the tables into
    config_data : dict
        Dict of tables to write
    html_tables: list
        List of dict for the tables extracted from the HTML
    soup: BeautifulSoup
        HTML page
    """

    for cell_pos, value in config_data.items():
        cell_x, cell_y = cell_pos

        if type(value) in [str, float, int]:
            value_to_write = value

        elif type(value) is list:
            value_to_write = []

            for (
                table_name_in_config,
                row,
                column,
                formatting,
            ) in value:
                subvalue = get_table_value_in_html_table(
                    table_name_in_config,
                    row,
                    column,
                    html_tables,
                    formatting,
                )
                value_to_write.append(subvalue)

            value_to_write = " ".join(value_to_write)

        # single value to add in the table
        elif type(value) is tuple:
            table_name_in_config, row, column = value
            value_to_write = get_table_value_in_html_table(
                table_name_in_config, row, column, html_tables
            )
        else:
            # special hardcoded case, haven't found a way to make that
            # better for now (which means it'll probably stay that way
            # forever)
            value_to_write = value(
                soup,
                "b",
                (
                    "Total number of somatic non-synonymous small "
                    "variants per megabase"
                ),
            )

        sheet.cell(cell_x, cell_y).value = value_to_write


def align_cells(sheet: Worksheet, config_data: list):
    """For given list of cells, align the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to align the cells
    config_data : list
        List of cells to align
    """

    for cell in config_data:
        sheet[cell].alignment = Alignment(wrapText=True, horizontal="center")


def bold_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells, bold them

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to bold the cells
    config_data : list
        List of cells to bold
    """

    for cell in config_data:
        sheet[cell].font = Font(bold=True, name=DEFAULT_FONT.name)


def set_col_width(sheet: Worksheet, config_data: list):
    """Given a list of columns, set their width

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to set the width
    config_data : list
        List of tuple with the column and its width to set
    """

    for cell, width in config_data:
        sheet.column_dimensions[cell].width = width


def color_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells and their color, color the cells appropriately

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to color the cells in
    config_data : list
        List of tuples with the cells and their color
    """

    for cell, color in config_data:
        sheet[cell].fill = color


def draw_borders(sheet: Worksheet, config_data: dict):
    """Draw borders around the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to draw borders
    config_data : dict
        Dict containing info for the single cells to draw borders around and
        the rows of cells
    """

    if config_data.get("single_cells"):
        for cell, type_border in config_data["single_cells"]:
            sheet[cell].border = type_border

    if config_data.get("cell_rows"):
        for cell_range, type_border in config_data["cell_rows"]:
            for cells in sheet[cell_range]:
                for cell in cells:
                    cell.border = type_border


def generate_dropdowns(sheet: Worksheet, config_data: dict):
    """Write in the dropdown menus

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the dropdown menus
    config_data : dict
        Dict of data for the dropdown menus
    """

    for dropdown_info in config_data:
        for cells, options in dropdown_info["cells"].items():
            dropdown = DataValidation(
                type="list", formula1=options, allow_blank=True
            )
            dropdown.prompt = "Select from the list"
            dropdown.promptTitle = dropdown_info["title"]
            dropdown.showInputMessage = True
            dropdown.showErrorMessage = True
            sheet.add_data_validation(dropdown)

            for cell in cells:
                dropdown.add(sheet[cell])


def insert_images(sheet: Worksheet, config_data: dict, images: list):
    """Insert images in the given worksheet for that config file

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the images
    config_data : list
        List of image data
    images: list
        List of images extracted from the HTML file
    """

    for image_data in config_data:
        height, width = image_data["size"]
        image_pil_obj = Image.open(images[image_data["img_index"]])
        image = drawing.image.Image(image_pil_obj)
        image.height = height
        image.width = width
        image.anchor = image_data["cell"]
        sheet.add_image(image)


def add_databar_rule(sheet: Worksheet, range_cell: str):
    """Add a databar for the range of cells given

    Parameters
    ----------
    sheet : Worksheet
        Sheet to add the databar(s) to
    range_cell : str
        String in "COL#:COL#" format for position of databar(s)
    """

    sheet.conditional_formatting.add(
        range_cell,
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="FF3361",
        ),
    )
